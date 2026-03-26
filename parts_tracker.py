#!/usr/bin/env python3
"""
Parts Tracker v1.0
Tracks SolidWorks parts across jobs via Everything HTTP API + PRF Excel data.

SETUP:
  1. pip install PyQt6 openpyxl requests
  2. In Everything: Tools > Options > HTTP Server > Enable (port 8080)
  3. Run: python parts_tracker.py
"""

import os
import re
import sys
import sqlite3
import subprocess
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, List

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QTabWidget, QTableWidget, QTableWidgetItem, QLineEdit, QPushButton,
    QLabel, QComboBox, QProgressBar, QDialog, QFormLayout, QMessageBox,
    QSplitter, QListWidget, QListWidgetItem, QHeaderView, QGroupBox,
    QStatusBar, QAbstractItemView, QFrame, QGridLayout,
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QTimer, QFileSystemWatcher
from PyQt6.QtGui import QFont, QIcon, QPixmap, QPainter, QColor, QPen, QPolygonF
from PyQt6.QtCore import QPointF
import math

# ── App Icon ───────────────────────────────────────────────────────────────
def make_icon() -> QIcon:
    """Draw a hex-bolt icon programmatically — no external file needed."""
    size = 64
    px = QPixmap(size, size)
    px.fill(QColor(0, 0, 0, 0))  # transparent

    p = QPainter(px)
    p.setRenderHint(QPainter.RenderHint.Antialiasing)

    cx, cy, r = size / 2, size / 2, size / 2 - 2

    # Hexagon
    hex_pts = QPolygonF([
        QPointF(cx + r * math.cos(math.radians(60 * i - 30)),
                cy + r * math.sin(math.radians(60 * i - 30)))
        for i in range(6)
    ])
    p.setBrush(QColor("#1e1e2e"))
    p.setPen(QPen(QColor("#89b4fa"), 3))
    p.drawPolygon(hex_pts)

    # Inner circle cutout ring
    p.setBrush(QColor(0, 0, 0, 0))
    p.setPen(QPen(QColor("#89b4fa"), 2))
    p.drawEllipse(QPointF(cx, cy), r * 0.38, r * 0.38)

    # "PT" text
    font = QFont("Segoe UI", 16, QFont.Weight.Bold)
    p.setFont(font)
    p.setPen(QColor("#cdd6f4"))
    p.drawText(px.rect(), Qt.AlignmentFlag.AlignCenter, "PT")

    p.end()
    return QIcon(px)


# ── Constants ──────────────────────────────────────────────────────────────
JOBS_ROOT = r"Z:\FOXFAB_DATA\ENGINEERING\2 JOBS"
EVERYTHING_URL = "http://localhost:8080/"
DB_PATH = Path(os.environ["APPDATA"]) / "PartsTracker" / "parts.db"

CATEGORIES = {
    "003": "Top Level Assembly",
    "100": "Subassembly",
    "200": "Metal",
    "240": "Copper",
    "245": "Flexibar",
    "250": "Galvanized",
    "295": "Insulation Barrier",
}

PART_RE  = re.compile(r"^(\d{3})-(\d{5})\.(sldprt|sldasm)$", re.IGNORECASE)
JOB_RE   = re.compile(r"^(J\d{5})([\s\-].*)?$", re.IGNORECASE)
SUBJ_RE  = re.compile(r"^(J\d{5}-\d{2})$",      re.IGNORECASE)

# ── Dark Theme ─────────────────────────────────────────────────────────────
STYLE = """
QMainWindow, QWidget, QDialog {
    background-color: #1e1e2e;
    color: #cdd6f4;
    font-family: 'Segoe UI', sans-serif;
    font-size: 13px;
}
QTabWidget::pane { border: 1px solid #313244; }
QTabBar::tab {
    background: #181825; color: #a6adc8;
    padding: 8px 20px; border: 1px solid #313244;
    border-bottom: none; border-radius: 4px 4px 0 0;
}
QTabBar::tab:selected { background: #313244; color: #cdd6f4; }
QTabBar::tab:hover    { background: #45475a; }

QTableWidget {
    background: #181825; alternate-background-color: #1e1e2e;
    gridline-color: #313244; border: 1px solid #313244;
    border-radius: 4px; selection-background-color: #45475a;
}
QTableWidget::item { padding: 4px 8px; }
QTableWidget::item:selected { background: #89b4fa; color: #1e1e2e; }
QHeaderView::section {
    background: #313244; color: #cdd6f4; padding: 6px 8px;
    border: none; border-right: 1px solid #45475a; font-weight: bold;
}

QLineEdit {
    background: #313244; color: #cdd6f4;
    border: 1px solid #45475a; border-radius: 6px; padding: 6px 10px;
}
QLineEdit:focus { border: 1px solid #89b4fa; }
QLineEdit:disabled { background: #1e1e2e; color: #585b70; }

QPushButton {
    background: #313244; color: #cdd6f4;
    border: 1px solid #45475a; border-radius: 6px; padding: 6px 14px;
}
QPushButton:hover   { background: #45475a; }
QPushButton:pressed { background: #585b70; }
QPushButton:disabled { color: #585b70; }
QPushButton#primary {
    background: #89b4fa; color: #1e1e2e;
    font-weight: bold; border: none;
}
QPushButton#primary:hover { background: #74c7ec; }
QPushButton#btn_open {
    background: #a6e3a1; color: #1e1e2e;
    font-weight: bold; border: none;
    padding: 3px 10px; border-radius: 4px;
}
QPushButton#btn_open:hover { background: #94e2d5; }
QPushButton#btn_open:disabled { background: #45475a; color: #585b70; }

QComboBox {
    background: #313244; color: #cdd6f4;
    border: 1px solid #45475a; border-radius: 6px; padding: 5px 10px;
}
QComboBox::drop-down { border: none; width: 20px; }
QComboBox QAbstractItemView {
    background: #313244; color: #cdd6f4;
    selection-background-color: #89b4fa; selection-color: #1e1e2e;
    border: 1px solid #45475a;
}

QProgressBar {
    background: #313244; border: 1px solid #45475a;
    border-radius: 4px; text-align: center; color: #cdd6f4;
}
QProgressBar::chunk { background: #89b4fa; border-radius: 4px; }

QListWidget {
    background: #181825; border: 1px solid #313244; border-radius: 4px;
}
QListWidget::item { padding: 8px 10px; border-bottom: 1px solid #313244; }
QListWidget::item:selected { background: #313244; color: #89b4fa; }
QListWidget::item:hover    { background: #24273a; }

QGroupBox {
    border: 1px solid #313244; border-radius: 6px;
    margin-top: 10px; padding-top: 10px;
    color: #a6adc8; font-weight: bold;
}
QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 5px; }

QStatusBar { background: #181825; color: #a6adc8; border-top: 1px solid #313244; }

QScrollBar:vertical {
    background: #1e1e2e; width: 8px; border-radius: 4px;
}
QScrollBar::handle:vertical {
    background: #45475a; border-radius: 4px; min-height: 20px;
}
QScrollBar::handle:vertical:hover { background: #585b70; }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }

QScrollBar:horizontal {
    background: #1e1e2e; height: 8px; border-radius: 4px;
}
QScrollBar::handle:horizontal {
    background: #45475a; border-radius: 4px; min-width: 20px;
}
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal { width: 0; }

QLabel#lbl_header  { font-size: 17px; font-weight: bold; color: #89b4fa; }
QLabel#lbl_sub     { font-size: 12px; color: #a6adc8; }
QLabel#lbl_badge   { font-size: 12px; color: #a6e3a1; }
QFrame#top_bar     { background: #181825; border-bottom: 1px solid #313244; }
QFrame#info_strip  { background: #181825; border: 1px solid #313244; border-radius: 6px; }
"""

# ── Database ───────────────────────────────────────────────────────────────
class Database:
    def __init__(self):
        DB_PATH.parent.mkdir(parents=True, exist_ok=True)
        self.con = sqlite3.connect(str(DB_PATH), check_same_thread=False)
        self.con.row_factory = sqlite3.Row
        self._init()

    def _init(self):
        self.con.executescript("""
            CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT);

            CREATE TABLE IF NOT EXISTS jobs (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                job_number      TEXT NOT NULL,
                job_name        TEXT,
                sub_job         TEXT NOT NULL,
                catalog_no      TEXT,
                enclosure_size  TEXT,
                prf_path        TEXT,
                scanned_at      TEXT,
                is_archived     INTEGER DEFAULT 0,
                UNIQUE(job_number, sub_job)
            );

            CREATE TABLE IF NOT EXISTS parts (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                part_number   TEXT NOT NULL,
                category_code TEXT,
                category_name TEXT,
                user_prefix   TEXT,
                file_ext      TEXT,
                full_path     TEXT NOT NULL UNIQUE,
                job_id        INTEGER REFERENCES jobs(id)
            );

            CREATE INDEX IF NOT EXISTS idx_p_user ON parts(user_prefix);
            CREATE INDEX IF NOT EXISTS idx_p_job  ON parts(job_id);
            CREATE INDEX IF NOT EXISTS idx_p_cat  ON parts(category_code);
        """)
        self.con.commit()
        # Migration: add is_archived to existing databases
        try:
            self.con.execute("ALTER TABLE jobs ADD COLUMN is_archived INTEGER DEFAULT 0")
            self.con.commit()
        except Exception:
            pass  # column already exists

    # settings
    def get(self, key, default=None):
        r = self.con.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
        return r["value"] if r else default

    def put(self, key, value):
        self.con.execute("INSERT OR REPLACE INTO settings VALUES(?,?)", (key, str(value)))
        self.con.commit()

    # jobs
    def job_id(self, job_number, sub_job):
        r = self.con.execute(
            "SELECT id FROM jobs WHERE job_number=? AND sub_job=?", (job_number, sub_job)
        ).fetchone()
        return r["id"] if r else None

    def upsert_job(self, job_number, job_name, sub_job,
                   catalog_no=None, enclosure_size=None, prf_path=None, is_archived=0):
        self.con.execute("""
            INSERT INTO jobs(job_number,job_name,sub_job,catalog_no,enclosure_size,prf_path,scanned_at,is_archived)
            VALUES(?,?,?,?,?,?,?,?)
            ON CONFLICT(job_number,sub_job) DO UPDATE SET
                job_name       = excluded.job_name,
                catalog_no     = COALESCE(excluded.catalog_no,    catalog_no),
                enclosure_size = COALESCE(excluded.enclosure_size, enclosure_size),
                prf_path       = COALESCE(excluded.prf_path,      prf_path),
                scanned_at     = excluded.scanned_at,
                is_archived    = excluded.is_archived
        """, (job_number, job_name, sub_job, catalog_no, enclosure_size, prf_path,
              datetime.now().isoformat(), is_archived))
        self.con.commit()
        return self.job_id(job_number, sub_job)

    def upsert_part(self, part_number, category_code, category_name,
                    user_prefix, file_ext, full_path, job_id):
        self.con.execute("""
            INSERT INTO parts(part_number,category_code,category_name,
                              user_prefix,file_ext,full_path,job_id)
            VALUES(?,?,?,?,?,?,?)
            ON CONFLICT(full_path) DO UPDATE SET job_id=excluded.job_id
        """, (part_number, category_code, category_name,
              user_prefix, file_ext, full_path, job_id))
        self.con.commit()

    def get_jobs(self, search="", size_f="", cat_f=""):
        q = """SELECT j.*, COUNT(p.id) as part_count
               FROM jobs j LEFT JOIN parts p ON p.job_id=j.id WHERE 1=1"""
        args = []
        if search:
            q += " AND (j.job_number LIKE ? OR j.job_name LIKE ? OR j.sub_job LIKE ?)"
            args += [f"%{search}%"] * 3
        if size_f:
            q += " AND j.enclosure_size LIKE ?"
            args.append(f"%{size_f}%")
        if cat_f:
            q += " AND j.catalog_no LIKE ?"
            args.append(f"%{cat_f}%")
        q += " GROUP BY j.id ORDER BY j.job_number, j.sub_job"
        return self.con.execute(q, args).fetchall()

    def get_parts(self, search="", user_prefix="", category="", job_id=None, file_ext=""):
        q = """SELECT p.*, j.job_number, j.job_name, j.sub_job,
                      j.catalog_no, j.enclosure_size, j.is_archived
               FROM parts p LEFT JOIN jobs j ON p.job_id=j.id WHERE 1=1"""
        args = []
        if search:
            q += " AND (p.part_number LIKE ? OR j.job_number LIKE ? OR j.job_name LIKE ?)"
            args += [f"%{search}%"] * 3
        if user_prefix:
            q += " AND p.user_prefix LIKE ?"
            args.append(f"{user_prefix}%")
        if category:
            q += " AND p.category_code=?"
            args.append(category)
        if job_id is not None:
            q += " AND p.job_id=?"
            args.append(job_id)
        if file_ext:
            q += " AND p.file_ext=?"
            args.append(file_ext)
        q += " ORDER BY p.part_number"
        return self.con.execute(q, args).fetchall()

    def user_part_count_for_job(self, job_id, user_prefix):
        r = self.con.execute(
            "SELECT COUNT(*) c FROM parts WHERE job_id=? AND user_prefix LIKE ?",
            (job_id, f"{user_prefix}%")
        ).fetchone()
        return r["c"]

    def unique_sizes(self):
        return [r[0] for r in self.con.execute(
            "SELECT DISTINCT enclosure_size FROM jobs WHERE enclosure_size IS NOT NULL ORDER BY 1"
        ).fetchall()]

    def unique_catalogs(self):
        return [r[0] for r in self.con.execute(
            "SELECT DISTINCT catalog_no FROM jobs WHERE catalog_no IS NOT NULL ORDER BY 1"
        ).fetchall()]

    def get_cat_prefixes(self) -> Dict[str, str]:
        """Return per-category prefix overrides, e.g. {'240': '804'}."""
        result = {}
        for code in CATEGORIES:
            if code == "003":
                continue
            val = self.get(f"cat_prefix_{code}")
            if val:
                result[code] = val
        return result

    def set_cat_prefixes(self, cat_prefixes: Dict[str, str]):
        """Persist per-category prefix overrides.  Pass empty string to clear a key."""
        for code in CATEGORIES:
            if code == "003":
                continue
            val = cat_prefixes.get(code, "")
            if val:
                self.put(f"cat_prefix_{code}", val)
            else:
                self.con.execute("DELETE FROM settings WHERE key=?", (f"cat_prefix_{code}",))
        self.con.commit()

    @staticmethod
    def _prefix_range(user_prefix: str):
        """Return (lo, hi) integer range covering this user's 5-digit part IDs.
        e.g. prefix '9'  → (90000, 99999)
             prefix '51' → (51000, 51999)"""
        p = int(user_prefix)
        pad = 5 - len(user_prefix)
        lo = p * (10 ** pad)
        hi = (p + 1) * (10 ** pad) - 1
        return lo, hi

    def latest_by_category(self, user_prefix: str = "",
                            cat_prefixes: Dict[str, str] = None) -> List[Dict]:
        """Return the highest part number per category whose file still exists on disk.
        Each category uses its own prefix override when provided, otherwise falls back
        to user_prefix.  Uses a fresh connection for an up-to-date read."""
        if cat_prefixes is None:
            cat_prefixes = {}
        con = sqlite3.connect(str(DB_PATH), timeout=10)
        con.row_factory = sqlite3.Row

        result: Dict[str, dict] = {}
        for cat_code, cat_name in CATEGORIES.items():
            if cat_code == "003":
                continue
            pfx = cat_prefixes.get(cat_code, user_prefix)
            if pfx:
                lo, hi = self._prefix_range(pfx)
                rows = con.execute("""
                    SELECT part_number, full_path
                    FROM parts
                    WHERE category_code = ?
                      AND CAST(user_prefix AS INTEGER) BETWEEN ? AND ?
                    ORDER BY part_number DESC
                """, (cat_code, lo, hi)).fetchall()
            else:
                rows = con.execute("""
                    SELECT part_number, full_path
                    FROM parts
                    WHERE category_code = ?
                    ORDER BY part_number DESC
                """, (cat_code,)).fetchall()

            for row in rows:
                if Path(row["full_path"]).exists():
                    result[cat_code] = {
                        "category_code": cat_code,
                        "category_name": cat_name,
                        "latest_part":   row["part_number"],
                        "effective_pfx": pfx,
                    }
                    break   # found the latest existing file for this category

        con.close()
        return list(result.values())

    def gaps_by_category(self, user_prefix: str,
                         cat_prefixes: Dict[str, str] = None) -> Dict[str, List[str]]:
        """Find skipped/missing part numbers per category for this user.
        Each category uses its own prefix override when set.
        Returns {cat_code: [missing_part_number, ...]} for categories with gaps."""
        if not user_prefix:
            return {}
        if cat_prefixes is None:
            cat_prefixes = {}

        con = sqlite3.connect(str(DB_PATH), timeout=10)
        con.row_factory = sqlite3.Row

        result: Dict[str, List[str]] = {}
        for cat_code in CATEGORIES:
            if cat_code == "003":
                continue
            pfx = cat_prefixes.get(cat_code, user_prefix)
            lo, hi = self._prefix_range(pfx)

            rows = con.execute("""
                SELECT part_number FROM parts
                WHERE category_code = ?
                  AND CAST(user_prefix AS INTEGER) BETWEEN ? AND ?
                ORDER BY part_number
            """, (cat_code, lo, hi)).fetchall()

            if len(rows) < 2:
                continue

            nums = []
            for row in rows:
                try:
                    nums.append(int(row["part_number"].split("-")[1]))
                except Exception:
                    continue

            if len(nums) < 2:
                continue

            nums.sort()
            present = set(nums)
            gaps = [
                f"{cat_code}-{str(n).zfill(5)}"
                for n in range(nums[0] + 1, nums[-1])
                if n not in present
            ]
            if gaps:
                result[cat_code] = gaps

        con.close()
        return result

    def clear_all(self):
        self.con.executescript("DELETE FROM parts; DELETE FROM jobs;")
        self.con.commit()

    def close(self):
        self.con.close()


# ── Everything Search ──────────────────────────────────────────────────────
def _eq(session, query: str) -> List[Dict]:
    """Run one Everything HTTP query, return [{name, path}]."""
    r = session.get(EVERYTHING_URL, params={
        "s": query, "j": 1, "path_column": 1, "count": 200000,
    }, timeout=60)
    r.raise_for_status()
    out = []
    for item in r.json().get("results", []):
        name = item.get("name") or item.get("filename", "")
        path = item.get("path", "")
        if name and path:
            out.append({"name": name, "path": path})
    return out


def find_003_folders(user_prefix: str) -> List[Path]:
    """
    Search Everything for 003-[userPrefix]*.sldasm files under JOBS_ROOT.
    Each hit anchors a parts folder — return unique folder paths.
    e.g. user_prefix='90' searches for '003-90*.sldasm'
    """
    import requests
    session = requests.Session()
    try:
        hits = _eq(session, f'"003-{user_prefix}" ext:sldasm path:"{JOBS_ROOT}"')
        if not hits:
            raise RuntimeError(
                f"No '003-{user_prefix}*.sldasm' files found.\n\n"
                f"Searched in: {JOBS_ROOT}\n\n"
                "Check that:\n"
                "1. Everything HTTP Server is enabled\n"
                "   (Everything → Tools → Options → HTTP Server, port 8080)\n"
                "2. Everything has indexed the Z:\\ drive\n"
                "3. Your user prefix is correct (current: " + user_prefix + ")"
            )
        folders = list({Path(h["path"]) for h in hits})
        return folders
    except requests.exceptions.ConnectionError:
        raise ConnectionError(
            "Cannot connect to Everything HTTP server.\n\n"
            "Enable it in Everything:\n"
            "Tools → Options → HTTP Server → Enable (port 8080)"
        )


# ── Gap Finder (Everything-based) ──────────────────────────────────────────
def find_gaps_via_everything(
    user_prefix: str,
    cat_prefixes: Dict[str, str] = None,
) -> Dict[str, List[str]]:
    """Compute genuine gaps by asking Everything for every file on disk.
    For each category, fetches ALL files matching the user's prefix under
    JOBS_ROOT, then finds numbers missing between the lowest and highest found.
    Because it queries Everything directly, parts that exist in ANY folder
    (scanned or not) are never reported as gaps.
    Returns {} on error (Everything unreachable, etc.).
    """
    if not user_prefix:
        return {}
    if cat_prefixes is None:
        cat_prefixes = {}
    try:
        import requests
        session = requests.Session()
        result: Dict[str, List[str]] = {}

        for cat_code in CATEGORIES:
            if cat_code == "003":
                continue
            pfx = cat_prefixes.get(cat_code, user_prefix)
            lo, hi = Database._prefix_range(pfx)

            # Fetch every matching file from Everything for this category+prefix
            hits: List[Dict] = []
            for ext in ("sldprt", "sldasm"):
                hits += _eq(
                    session,
                    f'"{cat_code}-{pfx}" ext:{ext} path:"{JOBS_ROOT}"',
                )

            # Collect numeric part of each matching filename, filtered to user range
            present: set = set()
            for hit in hits:
                m = PART_RE.match(hit["name"])
                if m and m.group(1) == cat_code:
                    n = int(m.group(2))
                    if lo <= n <= hi:
                        present.add(n)

            if len(present) < 2:
                continue

            min_n, max_n = min(present), max(present)
            gaps = [
                f"{cat_code}-{str(n).zfill(5)}"
                for n in range(min_n + 1, max_n)
                if n not in present
            ]
            if gaps:
                result[cat_code] = gaps

        return result

    except Exception:
        return {}   # Everything unreachable or error — show no gaps


# ── PRF Reader ─────────────────────────────────────────────────────────────
def find_prf(job_root: Path, sub_job: str) -> Optional[str]:
    prf_dir = job_root / "300 Inputs" / "302 Production Release Form"
    if not prf_dir.exists():
        return None
    # prefer file matching sub_job name
    sub_lower = sub_job.lower()
    candidates = [f for f in prf_dir.iterdir()
                  if f.suffix.lower() == ".xlsx" and "prf" in f.stem.lower()]
    for f in candidates:
        if sub_lower in f.stem.lower():
            return str(f)
    return str(candidates[0]) if candidates else None


def read_prf(path: str) -> Dict:
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        catalog = str(ws["C6"].value or "").strip() or None
        size_raw = str(ws["G18"].value or "").strip()
        m = re.search(r"(\d+)\s*[xX×]\s*(\d+)\s*[xX×]\s*(\d+)", size_raw)
        size = f"{m.group(1)} x {m.group(2)} x {m.group(3)}" if m else (size_raw or None)
        wb.close()
        return {"catalog_no": catalog, "enclosure_size": size}
    except Exception:
        return {"catalog_no": None, "enclosure_size": None}


# ── Path Parser ────────────────────────────────────────────────────────────
def parse_result(name: str, path: str) -> Optional[Dict]:
    m = PART_RE.match(name)
    if not m:
        return None
    cat_code   = m.group(1)
    user_pfx   = m.group(2)
    part_seq   = m.group(3)
    file_ext   = m.group(4).lower()
    part_num   = f"{cat_code}-{user_pfx}{part_seq}"
    cat_name   = CATEGORIES.get(cat_code, "Unknown")

    # Walk path parts to find job folder and sub-job folder
    parts_list = list(Path(path).parts)
    job_number = job_name_str = sub_job = None
    job_root_path = None

    for i, part in enumerate(parts_list):
        if SUBJ_RE.match(part):
            sub_job = part.upper()
        elif JOB_RE.match(part) and not SUBJ_RE.match(part):
            jm = JOB_RE.match(part)
            job_number = jm.group(1).upper()
            rest = (jm.group(2) or "").strip(" -").strip()
            job_name_str = rest or None
            job_root_path = Path(*parts_list[: i + 1])

    if not job_number or not sub_job:
        return None

    return {
        "name":        name,
        "part_number": part_num,
        "cat_code":    cat_code,
        "cat_name":    cat_name,
        "user_pfx":    user_pfx,
        "file_ext":    file_ext,
        "full_path":   os.path.join(path, name),
        "job_number":  job_number,
        "job_name":    job_name_str,
        "sub_job":     sub_job,
        "job_root":    job_root_path,
    }


# ── Scanner Thread ─────────────────────────────────────────────────────────
class ScanWorker(QThread):
    progress      = pyqtSignal(int, str)   # percent, message
    done          = pyqtSignal(int, int)   # new_parts, new_jobs
    err           = pyqtSignal(str)
    folders_found = pyqtSignal(list)       # list of folder path strings to watch

    def __init__(self, db: Database, user_prefix: str):
        super().__init__()
        self.db = db
        self.user_prefix = user_prefix

    def run(self):
        try:
            # Step 1: find folders containing user's 003- assembly
            self.progress.emit(2, f"Searching for 003-{self.user_prefix}* assemblies…")
            folders = find_003_folders(self.user_prefix)
            total = len(folders)

            self.progress.emit(5, f"Found {total} part folder(s) — processing…")

            new_parts = new_jobs = 0

            for i, folder in enumerate(folders):
                pct = 5 + int(90 * i / total)

                # Step 2: parse job info from the folder path.
                # Typical structure:
                #   Z:\FOXFAB_DATA\...\J15302 Garner Road\200 Mech\J15302-01\201 CAD
                # If the 003- file was moved to an archive sub-folder, Everything will
                # have returned that archive folder as the path — detect and flag it.
                parts_list = list(folder.parts)
                job_number = job_name_str = sub_job = None
                job_root_path = None

                for idx, part in enumerate(parts_list):
                    if SUBJ_RE.match(part):
                        sub_job = part.upper()
                    elif JOB_RE.match(part) and not SUBJ_RE.match(part):
                        jm = JOB_RE.match(part)
                        job_number = jm.group(1).upper()
                        rest = (jm.group(2) or "").strip(" -").strip()
                        job_name_str = rest or None
                        job_root_path = Path(*parts_list[: idx + 1])

                if not job_number:
                    continue
                # If no sub-job folder found, use the job number itself
                if not sub_job:
                    sub_job = job_number

                # Step 2a: verify this sub-job's folder actually contains the user's
                # 003- assembly at the top level (confirms it's the right sub-job).
                # find_003_folders already guarantees this for the returned folder, but
                # if the file is inside an "archive" sub-folder, flag the job.
                is_archived = any(p.lower() == "archive" for p in parts_list)

                key = (job_number, sub_job)

                # Step 3: upsert job — read PRF only if new
                existing_id = self.db.job_id(*key)
                if existing_id:
                    job_id = existing_id
                    # Always update the archived flag in case it changed
                    self.db.con.execute(
                        "UPDATE jobs SET is_archived=? WHERE id=?",
                        (1 if is_archived else 0, existing_id)
                    )
                    self.db.con.commit()
                else:
                    self.progress.emit(pct, f"New job {job_number} — reading PRF…")
                    prf_data = {"catalog_no": None, "enclosure_size": None}
                    prf_path = None
                    if job_root_path and job_root_path.exists():
                        prf_path = find_prf(job_root_path, sub_job)
                        if prf_path:
                            prf_data = read_prf(prf_path)
                    job_id = self.db.upsert_job(
                        job_number, job_name_str, sub_job,
                        prf_data["catalog_no"], prf_data["enclosure_size"],
                        prf_path, is_archived=1 if is_archived else 0
                    )
                    new_jobs += 1

                # Step 4: enumerate all sldprt/sldasm in this folder.
                # Only scan the immediate folder (no recursion) — archive sub-folders
                # are handled by the is_archived flag already set above.
                self.progress.emit(pct, f"Scanning {job_number}/{sub_job}…")
                try:
                    files = [f for f in folder.iterdir()
                             if f.is_file() and f.suffix.lower() in (".sldprt", ".sldasm")]
                except Exception:
                    continue

                for f in files:
                    m = PART_RE.match(f.name)
                    if not m:
                        continue
                    cat_code = m.group(1)
                    five_dig = m.group(2)   # 5-digit ID, e.g. "90015"
                    file_ext = m.group(3).lower()
                    part_num = f"{cat_code}-{five_dig}"
                    cat_name = CATEGORIES.get(cat_code, "Unknown")
                    self.db.upsert_part(
                        part_num, cat_code, cat_name,
                        five_dig, file_ext, str(f), job_id
                    )
                    new_parts += 1

            self.progress.emit(100, "Complete")
            self.folders_found.emit([str(f) for f in folders])
            self.done.emit(new_parts, new_jobs)

        except Exception as e:
            self.err.emit(str(e))


# ── Directory Re-scan Worker ───────────────────────────────────────────────
class DirectoryRescanWorker(QThread):
    """Lightweight re-scan of a single directory triggered by QFileSystemWatcher."""
    done = pyqtSignal()

    def __init__(self, db: Database, directory: str):
        super().__init__()
        self.db        = db
        self.directory = Path(directory)

    def run(self):
        try:
            dir_str = str(self.directory)

            # Find the job_id this directory belongs to via existing DB records
            row = self.db.con.execute(
                "SELECT job_id FROM parts WHERE full_path LIKE ? LIMIT 1",
                (dir_str + "%",)
            ).fetchone()
            if not row:
                return
            job_id = row["job_id"]

            # Enumerate current files on disk
            try:
                disk_files = {
                    str(f): f
                    for f in self.directory.iterdir()
                    if f.suffix.lower() in (".sldprt", ".sldasm") and PART_RE.match(f.name)
                }
            except Exception:
                return

            # Upsert every file currently on disk
            for path_str, f in disk_files.items():
                m = PART_RE.match(f.name)
                cat_code = m.group(1)
                five_dig = m.group(2)
                file_ext = m.group(3).lower()
                self.db.upsert_part(
                    f"{cat_code}-{five_dig}", cat_code,
                    CATEGORIES.get(cat_code, "Unknown"),
                    five_dig, file_ext, path_str, job_id
                )

            # Delete DB records for files that no longer exist on disk
            db_rows = self.db.con.execute(
                "SELECT id, full_path FROM parts WHERE full_path LIKE ?",
                (dir_str + "%",)
            ).fetchall()
            missing = [r["id"] for r in db_rows if r["full_path"] not in disk_files]
            if missing:
                self.db.con.executemany("DELETE FROM parts WHERE id=?", [(i,) for i in missing])
                self.db.con.commit()

            self.done.emit()
        except Exception:
            pass


# ── Scan Progress Dialog ───────────────────────────────────────────────────
class ScanDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Scanning…")
        self.setFixedSize(480, 130)
        self.setModal(True)
        self.setWindowFlags(Qt.WindowType.Dialog | Qt.WindowType.CustomizeWindowHint |
                            Qt.WindowType.WindowTitleHint)
        ly = QVBoxLayout(self)
        ly.setContentsMargins(24, 20, 24, 20)
        ly.setSpacing(10)

        self.lbl = QLabel("Initializing…")
        self.lbl.setObjectName("lbl_sub")
        ly.addWidget(self.lbl)

        self.bar = QProgressBar()
        self.bar.setRange(0, 100)
        self.bar.setFixedHeight(18)
        ly.addWidget(self.bar)

        self.detail = QLabel("")
        self.detail.setObjectName("lbl_sub")
        ly.addWidget(self.detail)

    def update(self, pct: int, msg: str):
        self.bar.setValue(pct)
        self.lbl.setText(msg)


# ── Setup Dialog ───────────────────────────────────────────────────────────
class SetupDialog(QDialog):
    def __init__(self, parent=None, prefill_id="", prefill_cats=None):
        super().__init__(parent)
        self.setWindowTitle("Parts Tracker — Setup")
        self.setMinimumWidth(460)
        self.setModal(True)

        self._prefill_cats = prefill_cats or {}
        self._cat_inputs: Dict[str, QLineEdit] = {}

        ly = QVBoxLayout(self)
        ly.setContentsMargins(36, 28, 36, 28)
        ly.setSpacing(14)

        title = QLabel("User Setup")
        title.setObjectName("lbl_header")
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        ly.addWidget(title)

        sub = QLabel("Enter your SolidWorks user ID.")
        sub.setObjectName("lbl_sub")
        sub.setAlignment(Qt.AlignmentFlag.AlignCenter)
        ly.addWidget(sub)

        # ── Main ID form ──
        form = QFormLayout()
        form.setSpacing(10)
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)

        self.id_inp = QLineEdit(prefill_id)
        self.id_inp.setPlaceholderText("e.g.  9  or  51  or  801")
        self.id_inp.setMaxLength(4)
        form.addRow("User ID:", self.id_inp)

        self.preview = QLabel("")
        self.preview.setObjectName("lbl_badge")
        form.addRow("Search prefix:", self.preview)
        self.id_inp.textChanged.connect(self._update_preview)
        self._update_preview()
        ly.addLayout(form)

        # ── More Options toggle ──
        self._more_btn = QPushButton("▶   More Options  —  per-category ID overrides")
        self._more_btn.setCheckable(True)
        self._more_btn.setChecked(False)
        self._more_btn.setStyleSheet(
            "QPushButton { text-align:left; padding:6px 10px;"
            " background:#24273a; border:1px solid #45475a; border-radius:6px;"
            " color:#a6adc8; font-size:12px; }"
            "QPushButton:hover { background:#313244; }"
            "QPushButton:checked { color:#89b4fa; border-color:#89b4fa; }"
        )
        self._more_btn.clicked.connect(self._toggle_more)
        ly.addWidget(self._more_btn)

        # ── Expandable section ──
        self._more_widget = QWidget()
        more_ly = QVBoxLayout(self._more_widget)
        more_ly.setContentsMargins(4, 0, 4, 0)
        more_ly.setSpacing(8)

        note = QLabel(
            "If you were assigned a different ID for specific categories, enter it here.\n"
            "Leave blank to use the main User ID above."
        )
        note.setObjectName("lbl_sub")
        note.setWordWrap(True)
        more_ly.addWidget(note)

        cat_form = QFormLayout()
        cat_form.setSpacing(8)
        cat_form.setLabelAlignment(Qt.AlignmentFlag.AlignRight)
        for code, name in CATEGORIES.items():
            if code == "003":
                continue
            inp = QLineEdit(self._prefill_cats.get(code, ""))
            inp.setMaxLength(4)
            inp.setPlaceholderText("same as main")
            cat_form.addRow(f"{code}  {name}:", inp)
            self._cat_inputs[code] = inp
        more_ly.addLayout(cat_form)

        self._more_widget.setVisible(False)
        ly.addWidget(self._more_widget)

        # ── Save button ──
        btn = QPushButton("Save && Continue")
        btn.setObjectName("primary")
        btn.clicked.connect(self._ok)
        ly.addWidget(btn)

        # Expand automatically if any overrides are already saved
        if any(self._prefill_cats.get(c) for c in self._cat_inputs):
            self._more_btn.setChecked(True)
            self._toggle_more(True)

        self.adjustSize()

    def _toggle_more(self, checked: bool):
        self._more_btn.setText(
            "▼   More Options  —  per-category ID overrides" if checked
            else "▶   More Options  —  per-category ID overrides"
        )
        self._more_widget.setVisible(checked)
        self.adjustSize()

    def _update_preview(self):
        uid = self.id_inp.text().strip()
        pad = "X" * (5 - len(uid))
        self.preview.setText(
            f"###-{uid}{pad}   (e.g. 003-{uid}{pad})" if uid.isdigit() else ""
        )
        # Update placeholders to reflect the current main ID
        for inp in self._cat_inputs.values():
            inp.setPlaceholderText(f"default: {uid}" if uid.isdigit() else "same as main")

    def _ok(self):
        uid = self.id_inp.text().strip()
        if not uid.isdigit():
            QMessageBox.warning(self, "Invalid", "User ID must be a number.")
            return
        # Validate any category overrides entered
        for code, inp in self._cat_inputs.items():
            v = inp.text().strip()
            if v and not v.isdigit():
                QMessageBox.warning(
                    self, "Invalid",
                    f"Override for category {code} must be a number (or leave blank)."
                )
                return
        self.accept()

    def values(self):
        uid = self.id_inp.text().strip()
        cat_prefixes = {}
        for code, inp in self._cat_inputs.items():
            v = inp.text().strip()
            if v and v.isdigit():
                cat_prefixes[code] = v
        return {"user_id": uid, "user_prefix": uid, "cat_prefixes": cat_prefixes}


# ── Parts Table ────────────────────────────────────────────────────────────
PART_COLS = ["Part Number", "Category", "Type", "Job", "Sub-Job",
             "Catalog No", "Enclosure Size", ""]

class PartsTable(QTableWidget):
    open_clicked = pyqtSignal(str)

    def __init__(self):
        super().__init__()
        self.setColumnCount(len(PART_COLS))
        self.setHorizontalHeaderLabels(PART_COLS)
        self.setAlternatingRowColors(True)
        self.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.setShowGrid(False)
        self.verticalHeader().setVisible(False)
        hh = self.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(7, QHeaderView.ResizeMode.Fixed)
        self.setColumnWidth(7, 70)

    def load(self, rows):
        self.setUpdatesEnabled(False)
        self.setRowCount(0)
        _archived_bg = QColor("#2a2000")
        _archived_fg = QColor("#f9e2af")
        for r, p in enumerate(rows):
            self.insertRow(r)
            self.setRowHeight(r, 30)
            # Highlight only if THIS specific file lives inside an "archive" folder
            archived = any(
                part.lower() == "archive"
                for part in Path(p["full_path"]).parts
            )
            job_str = f"{p['job_number']} {p['job_name'] or ''}".strip()
            vals = [
                p["part_number"],
                f"{p['category_code']} – {p['category_name']}",
                p["file_ext"].upper(),
                job_str,
                p["sub_job"] or "—",
                p["catalog_no"] or "—",
                p["enclosure_size"] or "—",
            ]
            for c, v in enumerate(vals):
                item = QTableWidgetItem(str(v))
                item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
                if archived:
                    item.setBackground(_archived_bg)
                    item.setForeground(_archived_fg)
                self.setItem(r, c, item)
            btn = QPushButton("Open")
            btn.setObjectName("btn_open")
            btn.setFixedHeight(24)
            fp = p["full_path"]
            btn.clicked.connect(lambda _, x=fp: self.open_clicked.emit(x))
            self.setCellWidget(r, 7, btn)
        self.setUpdatesEnabled(True)


# ── My Parts Tab ───────────────────────────────────────────────────────────
class MyPartsTab(QWidget):
    def __init__(self, db: Database, user_prefix: str):
        super().__init__()
        self.db = db
        self.user_prefix = user_prefix
        self._cur_job_id = None
        self._build()

    def _build(self):
        ly = QHBoxLayout(self)
        ly.setContentsMargins(10, 10, 10, 10)
        ly.setSpacing(0)

        splitter = QSplitter(Qt.Orientation.Horizontal)

        # ── Left sidebar ──
        left = QWidget()
        ll = QVBoxLayout(left)
        ll.setContentsMargins(0, 0, 8, 0)
        ll.setSpacing(6)

        lbl = QLabel("Jobs")
        lbl.setObjectName("lbl_header")
        ll.addWidget(lbl)

        self.job_search = QLineEdit()
        self.job_search.setPlaceholderText("Filter jobs…")
        self.job_search.textChanged.connect(self._load_jobs)
        ll.addWidget(self.job_search)

        self.job_list = QListWidget()
        self.job_list.currentItemChanged.connect(self._job_selected)
        ll.addWidget(self.job_list)

        self.job_count_lbl = QLabel("")
        self.job_count_lbl.setObjectName("lbl_sub")
        ll.addWidget(self.job_count_lbl)

        left.setMinimumWidth(220)
        left.setMaximumWidth(320)
        splitter.addWidget(left)

        # ── Right panel ──
        right = QWidget()
        rl = QVBoxLayout(right)
        rl.setContentsMargins(8, 0, 0, 0)
        rl.setSpacing(6)

        # Header row
        hr = QHBoxLayout()
        self.parts_title = QLabel("Select a job")
        self.parts_title.setObjectName("lbl_header")
        hr.addWidget(self.parts_title)
        hr.addStretch()

        self.cat_cb = QComboBox()
        self.cat_cb.addItem("All Categories", "")
        for code, name in CATEGORIES.items():
            self.cat_cb.addItem(f"{code} – {name}", code)
        self.cat_cb.currentIndexChanged.connect(self._load_parts)
        hr.addWidget(self.cat_cb)

        self.type_cb = QComboBox()
        self.type_cb.addItems(["All Types", "sldprt", "sldasm"])
        self.type_cb.currentIndexChanged.connect(self._load_parts)
        hr.addWidget(self.type_cb)
        rl.addLayout(hr)

        # Info strip
        self.info_strip = QFrame()
        self.info_strip.setObjectName("info_strip")
        self.info_strip.setFixedHeight(36)
        isl = QHBoxLayout(self.info_strip)
        isl.setContentsMargins(12, 0, 12, 0)
        self.info_lbl = QLabel("")
        self.info_lbl.setObjectName("lbl_sub")
        isl.addWidget(self.info_lbl)
        isl.addStretch()
        self.part_count_lbl = QLabel("")
        self.part_count_lbl.setObjectName("lbl_badge")
        isl.addWidget(self.part_count_lbl)
        rl.addWidget(self.info_strip)

        self.table = PartsTable()
        self.table.open_clicked.connect(open_path)
        rl.addWidget(self.table)

        splitter.addWidget(right)
        splitter.setStretchFactor(0, 0)
        splitter.setStretchFactor(1, 1)
        ly.addWidget(splitter)
        self._load_jobs()

    def _load_jobs(self):
        search = self.job_search.text().strip()
        all_jobs = self.db.get_jobs(search=search)
        self.job_list.clear()
        user_jobs = []
        for job in all_jobs:
            if job["part_count"] == 0:
                continue
            # Sort by the 003- assembly number in this job
            row = self.db.con.execute(
                "SELECT part_number FROM parts WHERE job_id=? AND category_code='003' LIMIT 1",
                (job["id"],)
            ).fetchone()
            sort_key = row["part_number"] if row else "003-99999"
            user_jobs.append((sort_key, dict(job)))

        user_jobs.sort(key=lambda x: x[0])

        for _, job in user_jobs:
            text = f"{job['sub_job']}\n{job['job_name'] or ''}"
            item = QListWidgetItem(text)
            item.setData(Qt.ItemDataRole.UserRole, job)
            self.job_list.addItem(item)
        self.job_count_lbl.setText(f"{len(user_jobs)} job(s)")

    def _job_selected(self, item):
        if not item:
            return
        data = item.data(Qt.ItemDataRole.UserRole)
        self._cur_job_id = data["id"]
        self.parts_title.setText(f"{data['sub_job']}  —  {data['job_name'] or ''}")
        info = []
        if data.get("catalog_no"):     info.append(f"Catalog: {data['catalog_no']}")
        if data.get("enclosure_size"): info.append(f"Size: {data['enclosure_size']}")
        if data.get("is_archived"):    info.append("⚠ Parts Archived")
        self.info_lbl.setText("   ·   ".join(info) if info else "No PRF data found")
        # Colour the archived warning in amber
        if data.get("is_archived"):
            self.info_lbl.setStyleSheet("color:#f9e2af; font-size:12px;")
        else:
            self.info_lbl.setStyleSheet("")
        self._load_parts()

    def _load_parts(self):
        if self._cur_job_id is None:
            return
        cat  = self.cat_cb.currentData()
        ext  = self.type_cb.currentText() if self.type_cb.currentIndex() > 0 else ""
        rows = self.db.get_parts(category=cat or "",
                                  job_id=self._cur_job_id, file_ext=ext)
        self.table.load(rows)
        self.part_count_lbl.setText(f"{len(rows)} part(s)")

    def refresh(self, user_prefix: str = None):
        if user_prefix:
            self.user_prefix = user_prefix
        self._load_jobs()
        if self._cur_job_id:
            self._load_parts()


# ── All Parts Tab ──────────────────────────────────────────────────────────
class AllPartsTab(QWidget):
    def __init__(self, db: Database, user_prefix: str):
        super().__init__()
        self.db = db
        self.user_prefix = user_prefix
        self._build()

    def _build(self):
        ly = QVBoxLayout(self)
        ly.setContentsMargins(10, 10, 10, 10)
        ly.setSpacing(8)

        hr = QHBoxLayout()
        lbl = QLabel("All Parts")
        lbl.setObjectName("lbl_header")
        hr.addWidget(lbl)
        hr.addStretch()
        self.count_lbl = QLabel("")
        self.count_lbl.setObjectName("lbl_sub")
        hr.addWidget(self.count_lbl)
        ly.addLayout(hr)

        fg = QGroupBox("Filters")
        fl = QHBoxLayout(fg)
        fl.setSpacing(10)

        self.search_inp = QLineEdit()
        self.search_inp.setPlaceholderText("Search part number or job…")
        self.search_inp.textChanged.connect(self._refresh)
        fl.addWidget(self.search_inp, 2)

        fl.addWidget(QLabel("Category:"))
        self.cat_cb = QComboBox()
        self.cat_cb.addItem("All", "")
        for code, name in CATEGORIES.items():
            self.cat_cb.addItem(f"{code} – {name}", code)
        self.cat_cb.currentIndexChanged.connect(self._refresh)
        fl.addWidget(self.cat_cb)

        fl.addWidget(QLabel("Type:"))
        self.type_cb = QComboBox()
        self.type_cb.addItems(["All", "sldprt", "sldasm"])
        self.type_cb.currentIndexChanged.connect(self._refresh)
        fl.addWidget(self.type_cb)

        clr = QPushButton("Clear")
        clr.clicked.connect(self._clear)
        fl.addWidget(clr)
        ly.addWidget(fg)

        self.table = PartsTable()
        self.table.open_clicked.connect(open_path)
        ly.addWidget(self.table)

    def _clear(self):
        self.search_inp.clear()
        self.cat_cb.setCurrentIndex(0)
        self.type_cb.setCurrentIndex(0)

    def _refresh(self):
        search = self.search_inp.text().strip()
        cat    = self.cat_cb.currentData()
        ext    = self.type_cb.currentText() if self.type_cb.currentIndex() > 0 else ""
        rows   = self.db.get_parts(search=search, category=cat or "", file_ext=ext)
        self.table.load(rows)
        self.count_lbl.setText(f"{len(rows):,} result(s)")

    def refresh(self, user_prefix: str = None):
        if user_prefix:
            self.user_prefix = user_prefix
        self._refresh()


# ── Jobs Tab ───────────────────────────────────────────────────────────────
class JobsTab(QWidget):
    def __init__(self, db: Database):
        super().__init__()
        self.db = db
        self._build()

    def _build(self):
        ly = QVBoxLayout(self)
        ly.setContentsMargins(10, 10, 10, 10)
        ly.setSpacing(8)

        hr = QHBoxLayout()
        lbl = QLabel("Jobs")
        lbl.setObjectName("lbl_header")
        hr.addWidget(lbl)
        hr.addStretch()
        self.count_lbl = QLabel("")
        self.count_lbl.setObjectName("lbl_sub")
        hr.addWidget(self.count_lbl)
        ly.addLayout(hr)

        # Filters
        fg = QGroupBox("Filters")
        fl = QHBoxLayout(fg)
        fl.setSpacing(10)

        self.search_inp = QLineEdit()
        self.search_inp.setPlaceholderText("Search job number or name…")
        self.search_inp.textChanged.connect(self._refresh)
        fl.addWidget(self.search_inp, 2)

        fl.addWidget(QLabel("Enclosure Size:"))
        self.size_cb = QComboBox()
        self.size_cb.setEditable(True)
        self.size_cb.setMinimumWidth(160)
        self.size_cb.currentTextChanged.connect(self._refresh)
        fl.addWidget(self.size_cb)

        fl.addWidget(QLabel("Catalog:"))
        self.cat_cb = QComboBox()
        self.cat_cb.setEditable(True)
        self.cat_cb.setMinimumWidth(200)
        self.cat_cb.currentTextChanged.connect(self._refresh)
        fl.addWidget(self.cat_cb)

        clr = QPushButton("Clear")
        clr.clicked.connect(self._clear)
        fl.addWidget(clr)
        ly.addWidget(fg)

        # Jobs table
        JCOLS = ["Job #", "Sub-Job", "Job Name",
                  "Catalog No", "Enclosure Size", "Parts", "Scanned", ""]
        self.tbl = QTableWidget()
        self.tbl.setColumnCount(len(JCOLS))
        self.tbl.setHorizontalHeaderLabels(JCOLS)
        self.tbl.setAlternatingRowColors(True)
        self.tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.tbl.setShowGrid(False)
        self.tbl.verticalHeader().setVisible(False)
        hh = self.tbl.horizontalHeader()
        hh.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        hh.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        hh.setSectionResizeMode(7, QHeaderView.ResizeMode.Fixed)
        self.tbl.setColumnWidth(7, 70)
        ly.addWidget(self.tbl)

    def _clear(self):
        self.search_inp.clear()
        self.size_cb.setCurrentIndex(0)
        self.cat_cb.setCurrentIndex(0)

    def _refresh(self):
        search = self.search_inp.text().strip()
        size_t = self.size_cb.currentText()
        cat_t  = self.cat_cb.currentText()
        size_f = "" if size_t in ("", "All Sizes")    else size_t
        cat_f  = "" if cat_t  in ("", "All Catalogs") else cat_t
        jobs   = self.db.get_jobs(search=search, size_f=size_f, cat_f=cat_f)

        self.tbl.setUpdatesEnabled(False)
        self.tbl.setRowCount(0)
        for r, job in enumerate(jobs):
            self.tbl.insertRow(r)
            self.tbl.setRowHeight(r, 30)
            scanned = job["scanned_at"] or ""
            try:
                scanned = datetime.fromisoformat(scanned).strftime("%Y-%m-%d") if scanned else ""
            except Exception:
                pass
            archived = bool(job["is_archived"]) if "is_archived" in job.keys() else False
            job_name_display = job["job_name"] or ""
            if archived:
                job_name_display = f"{job_name_display}  [Archived]".strip()
            vals = [job["job_number"], job["sub_job"], job_name_display,
                    job["catalog_no"] or "—", job["enclosure_size"] or "—",
                    str(job["part_count"]), scanned]
            for c, v in enumerate(vals):
                item = QTableWidgetItem(str(v))
                item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
                if archived:
                    item.setForeground(QColor("#f9e2af"))
                self.tbl.setItem(r, c, item)
            # Open folder button
            btn = QPushButton("Open")
            btn.setObjectName("btn_open")
            btn.setFixedHeight(24)
            part_row = self.db.con.execute(
                "SELECT full_path FROM parts WHERE job_id=? LIMIT 1", (job["id"],)
            ).fetchone()
            if part_row:
                folder = str(Path(part_row["full_path"]).parent)
                btn.clicked.connect(lambda _, f=folder: open_path(f, folder=True))
            else:
                btn.setEnabled(False)
            self.tbl.setCellWidget(r, 7, btn)
        self.tbl.setUpdatesEnabled(True)
        self.count_lbl.setText(f"{len(jobs):,} job(s)")

    def refresh(self):
        sizes    = self.db.unique_sizes()
        catalogs = self.db.unique_catalogs()
        cur_sz   = self.size_cb.currentText()
        cur_cat  = self.cat_cb.currentText()

        self.size_cb.blockSignals(True)
        self.size_cb.clear()
        self.size_cb.addItem("All Sizes")
        self.size_cb.addItems(sizes)
        idx = self.size_cb.findText(cur_sz)
        self.size_cb.setCurrentIndex(max(0, idx))
        self.size_cb.blockSignals(False)

        self.cat_cb.blockSignals(True)
        self.cat_cb.clear()
        self.cat_cb.addItem("All Catalogs")
        self.cat_cb.addItems(catalogs)
        idx = self.cat_cb.findText(cur_cat)
        self.cat_cb.setCurrentIndex(max(0, idx))
        self.cat_cb.blockSignals(False)

        self._refresh()


# ── Next Numbers Tab ──────────────────────────────────────────────────────
# Category accent colours (subtle left-border on each card)
CAT_COLORS = {
    "100": "#89b4fa",  # blue   – Subassembly
    "200": "#a6adc8",  # grey   – Metal
    "240": "#fab387",  # orange – Copper
    "245": "#a6e3a1",  # green  – Flexibar
    "250": "#89dceb",  # cyan   – Galvanized
    "295": "#f9e2af",  # yellow – Insulation Barrier
}

CARD_STYLE = """
QFrame#part_card {{
    background: #24273a;
    border: 1px solid #313244;
    border-left: 4px solid {color};
    border-radius: 8px;
}}
"""


def _next_part(user_prefix: str, cat_code: str, latest: str) -> str:
    """Given the latest part number string, return the next one."""
    if latest:
        five = latest.split("-")[1]        # e.g. "90015"
        nxt  = str(int(five) + 1).zfill(5)
    else:
        # first ever part: prefix + zeros + 1
        zeros = "0" * (4 - len(user_prefix))
        nxt   = f"{user_prefix}{zeros}1"
    return f"{cat_code}-{nxt}"


class GapScanWorker(QThread):
    finished = pyqtSignal(dict)

    def __init__(self, user_prefix: str, cat_prefixes: Dict[str, str]):
        super().__init__()
        self.user_prefix  = user_prefix
        self.cat_prefixes = cat_prefixes

    def run(self):
        result = find_gaps_via_everything(self.user_prefix, self.cat_prefixes)
        self.finished.emit(result)


class NextNumbersTab(QWidget):
    def __init__(self, db: Database, user_prefix: str, cat_prefixes: Dict[str, str] = None):
        super().__init__()
        self.db           = db
        self.user_prefix  = user_prefix
        self.cat_prefixes = cat_prefixes or {}
        self._cards: Dict[str, dict] = {}
        self._cached_gaps = None
        self._gap_worker: Optional[GapScanWorker] = None
        self._build()

        self._poll_timer = QTimer(self)
        self._poll_timer.timeout.connect(self._safe_refresh)
        self._poll_timer.start(5_000)

        # Auto-scan gaps on startup
        if self.user_prefix:
            QTimer.singleShot(300, self._start_gap_scan)

    def _safe_refresh(self):
        try:
            self.refresh()
        except Exception:
            pass  # skip this tick on transient DB errors (e.g. lock during scan)

    def _build(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(16, 14, 16, 14)
        outer.setSpacing(12)

        # Header
        hdr = QHBoxLayout()
        title = QLabel("Next Part Numbers")
        title.setObjectName("lbl_header")
        hdr.addWidget(title)
        hdr.addStretch()
        self.sub_lbl = QLabel("Your next available number per category")
        self.sub_lbl.setObjectName("lbl_sub")
        hdr.addWidget(self.sub_lbl)
        outer.addLayout(hdr)

        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("color: #313244;")
        outer.addWidget(sep)

        # Grid of cards — 3 per row
        grid_widget = QWidget()
        self.grid = QGridLayout(grid_widget)
        self.grid.setSpacing(14)
        self.grid.setContentsMargins(0, 0, 0, 0)

        categories = [(k, v) for k, v in CATEGORIES.items() if k != "003"]
        for idx, (code, name) in enumerate(categories):
            color = CAT_COLORS.get(code, "#cdd6f4")
            card  = self._make_card(code, name, color)
            self.grid.addWidget(card, idx // 3, idx % 3)

        outer.addWidget(grid_widget)

        self._gap_panel = self._build_gap_section()
        outer.addWidget(self._gap_panel)
        outer.addStretch()

    def _build_gap_section(self) -> QFrame:
        panel = QFrame()
        panel.setObjectName("gap_panel")
        panel.setStyleSheet(
            "QFrame#gap_panel { background:#181825; border:1px solid #313244; border-radius:8px; }"
        )
        pl = QVBoxLayout(panel)
        pl.setContentsMargins(16, 14, 16, 14)
        pl.setSpacing(0)

        # Header row
        hdr = QHBoxLayout()
        hdr.setSpacing(10)
        title = QLabel("Gap Analysis")
        title.setStyleSheet("font-size:14px; font-weight:bold; color:#f9e2af;")
        hdr.addWidget(title)
        self._gap_status_lbl = QLabel("Pending — scan or open tab to check")
        self._gap_status_lbl.setStyleSheet("font-size:11px; color:#585b70;")
        hdr.addWidget(self._gap_status_lbl)
        hdr.addStretch()
        self._gap_scan_btn = QPushButton("Scan Gaps")
        self._gap_scan_btn.setStyleSheet(
            "QPushButton { background:#f9e2af; color:#1e1e2e; border:none; border-radius:5px;"
            " font-size:11px; font-weight:bold; padding:4px 12px; }"
            "QPushButton:hover { background:#fde68a; }"
            "QPushButton:disabled { background:#45475a; color:#585b70; }"
        )
        self._gap_scan_btn.clicked.connect(self._scan_gaps_now)
        hdr.addWidget(self._gap_scan_btn)
        pl.addLayout(hdr)

        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.HLine)
        sep.setStyleSheet("background:#313244; max-height:1px; margin-top:10px; margin-bottom:6px;")
        pl.addWidget(sep)

        # Per-category rows (hidden until gaps are loaded)
        self._gap_rows: Dict[str, dict] = {}
        for code, name in CATEGORIES.items():
            if code == "003":
                continue
            color = CAT_COLORS.get(code, "#cdd6f4")
            row_w = self._make_gap_row(code, name, color)
            pl.addWidget(row_w)

        return panel

    def _make_gap_row(self, code: str, name: str, color: str) -> QWidget:
        container = QWidget()
        cl = QVBoxLayout(container)
        cl.setContentsMargins(0, 1, 0, 1)
        cl.setSpacing(0)

        toggle = QPushButton(f"  {name}  ({code})")
        toggle.setFlat(True)
        toggle.setCursor(Qt.CursorShape.PointingHandCursor)
        cl.addWidget(toggle)

        detail = QLabel()
        detail.setWordWrap(True)
        detail.setVisible(False)
        detail.setStyleSheet(
            "color:#a6adc8; font-size:11px; padding:4px 6px 8px 28px;"
            " font-family:'Consolas','Courier New',monospace;"
        )
        cl.addWidget(detail)

        toggle.clicked.connect(lambda _, t=toggle, d=detail: self._toggle_gap_row(t, d))
        self._gap_rows[code] = {"container": container, "toggle": toggle, "detail": detail, "color": color}
        container.setVisible(False)
        return container

    def _toggle_gap_row(self, toggle: QPushButton, detail: QLabel):
        expanding = not detail.isVisible()
        detail.setVisible(expanding)
        txt = toggle.text()
        if expanding:
            toggle.setText(txt.replace("▶", "▼", 1))
        else:
            toggle.setText(txt.replace("▼", "▶", 1))

    def _start_gap_scan(self):
        if self._gap_worker and self._gap_worker.isRunning():
            return
        if not self.user_prefix:
            return
        self._gap_status_lbl.setText("Scanning Everything…")
        self._gap_status_lbl.setStyleSheet("font-size:11px; color:#89b4fa;")
        self._gap_scan_btn.setEnabled(False)
        worker = GapScanWorker(self.user_prefix, self.cat_prefixes)
        worker.finished.connect(self._on_gap_scan_done)
        worker.finished.connect(worker.deleteLater)
        self._gap_worker = worker
        worker.start()

    def _on_gap_scan_done(self, result: dict):
        self._cached_gaps = result
        self._gap_worker = None
        self._gap_scan_btn.setEnabled(True)
        self.refresh()

    def _scan_gaps_now(self):
        self._start_gap_scan()

    def _make_card(self, code: str, name: str, color: str) -> QFrame:
        card = QFrame()
        card.setObjectName("part_card")
        card.setStyleSheet(CARD_STYLE.format(color=color))
        card.setMinimumHeight(130)

        ly = QVBoxLayout(card)
        ly.setContentsMargins(16, 14, 16, 14)
        ly.setSpacing(6)

        # Category title
        cat_lbl = QLabel(f"{name}")
        cat_lbl.setStyleSheet(f"font-size:14px; font-weight:bold; color:{color};")
        ly.addWidget(cat_lbl)

        code_lbl = QLabel(f"Category {code}")
        code_lbl.setStyleSheet("font-size:11px; color:#585b70;")
        ly.addWidget(code_lbl)

        ly.addSpacing(4)

        # Latest row
        latest_row = QHBoxLayout()
        latest_row.setSpacing(8)
        latest_title = QLabel("Latest:")
        latest_title.setStyleSheet("color:#a6adc8; font-size:12px;")
        latest_title.setFixedWidth(50)
        self._cards.setdefault(code, {})
        latest_val = QLabel("—")
        latest_val.setStyleSheet("color:#cdd6f4; font-size:13px; font-weight:bold;")
        self._cards[code]["latest_lbl"] = latest_val
        latest_row.addWidget(latest_title)
        latest_row.addWidget(latest_val)
        latest_row.addStretch()
        ly.addLayout(latest_row)

        # Next row
        next_row = QHBoxLayout()
        next_row.setSpacing(8)
        next_title = QLabel("Next:")
        next_title.setStyleSheet("color:#a6adc8; font-size:12px;")
        next_title.setFixedWidth(50)
        next_val = QLabel("—")
        next_val.setStyleSheet(f"color:{color}; font-size:16px; font-weight:bold;")
        self._cards[code]["next_lbl"] = next_val
        self._cards[code]["code"]     = code

        gap_badge = QLabel("GAP")
        gap_badge.setStyleSheet(
            "background:#f9e2af; color:#1e1e2e; font-size:10px; font-weight:bold;"
            " border-radius:3px; padding:1px 5px;"
        )
        gap_badge.setVisible(False)
        self._cards[code]["gap_badge"] = gap_badge

        copy_btn = QPushButton("Copy")
        copy_btn.setFixedSize(52, 24)
        copy_btn.setStyleSheet(
            f"QPushButton {{ background:transparent; color:{color}; border:1px solid {color};"
            f" border-radius:4px; font-size:11px; padding:0; }}"
            f"QPushButton:hover {{ background:{color}; color:#1e1e2e; }}"
        )
        copy_btn.clicked.connect(lambda _, lbl=next_val, c=code: self._copy_next(lbl.text(), c))

        next_row.addWidget(next_title)
        next_row.addWidget(next_val)
        next_row.addSpacing(6)
        next_row.addWidget(gap_badge)
        next_row.addStretch()
        next_row.addWidget(copy_btn)
        ly.addLayout(next_row)

        return card

    def _copy_next(self, text: str, code: str):
        if not text or text == "—":
            return
        QApplication.clipboard().setText(text)
        # Remove used gap from cache so next gap (or latest+1) shows next
        if self._cached_gaps and code in self._cached_gaps:
            try:
                self._cached_gaps[code].remove(text)
            except ValueError:
                pass
            if not self._cached_gaps[code]:
                del self._cached_gaps[code]
        self.refresh()

    def refresh(self, user_prefix: str = None, cat_prefixes: Dict[str, str] = None):
        if user_prefix:
            self.user_prefix = user_prefix
        if cat_prefixes is not None:
            self.cat_prefixes = cat_prefixes

        rows = self.db.latest_by_category(self.user_prefix, self.cat_prefixes)
        existing = {r["category_code"]: r for r in rows}

        gaps = self._cached_gaps if self._cached_gaps is not None else {}

        for code in self._cards:
            row      = existing.get(code)
            latest   = row["latest_part"] if row else ""
            eff_pfx  = self.cat_prefixes.get(code, self.user_prefix)
            gap_list = gaps.get(code, [])

            if gap_list:
                nxt      = gap_list[0]  # lowest missing number first
                is_gap   = True
            else:
                nxt      = _next_part(eff_pfx, code, latest)
                is_gap   = False

            self._cards[code]["latest_lbl"].setText(latest if latest else "None yet")
            self._cards[code]["next_lbl"].setText(nxt)

            badge = self._cards[code].get("gap_badge")
            if badge:
                badge.setVisible(is_gap)

        # ── Update gap panel ──
        scanned = self._cached_gaps is not None
        cats_with_gaps = 0
        for code, row_info in self._gap_rows.items():
            cat_name  = CATEGORIES.get(code, code)
            color     = row_info["color"]
            gap_list  = gaps.get(code, [])
            toggle    = row_info["toggle"]
            detail    = row_info["detail"]
            container = row_info["container"]

            if not scanned:
                container.setVisible(False)
                continue

            # Preserve expanded/collapsed state across refreshes
            was_expanded = detail.isVisible()

            container.setVisible(True)
            if gap_list:
                cats_with_gaps += 1
                n     = len(gap_list)
                label = "gap" if n == 1 else "gaps"
                toggle.setStyleSheet(
                    f"QPushButton {{ background:transparent; color:{color}; border:none;"
                    f" font-size:12px; font-weight:bold; text-align:left; padding:5px 4px; }}"
                    f"QPushButton:hover {{ color:#cdd6f4; }}"
                )
                arrow = "▼" if was_expanded else "▶"
                toggle.setText(f"{arrow}  {cat_name}  ({code})   —   {n} {label} missing")
                detail.setText("  ".join(gap_list))
                detail.setVisible(was_expanded)
            else:
                toggle.setStyleSheet(
                    "QPushButton { background:transparent; color:#585b70; border:none;"
                    " font-size:12px; text-align:left; padding:5px 4px; }"
                )
                toggle.setText(f"  ✓  {cat_name}  ({code})   —   no gaps")
                detail.setVisible(False)

        if scanned:
            if cats_with_gaps:
                self._gap_status_lbl.setText(f"{cats_with_gaps} categor{'y' if cats_with_gaps == 1 else 'ies'} with gaps")
                self._gap_status_lbl.setStyleSheet("font-size:11px; color:#f9e2af;")
            else:
                self._gap_status_lbl.setText("All clear — no gaps found")
                self._gap_status_lbl.setStyleSheet("font-size:11px; color:#a6e3a1;")
        else:
            self._gap_status_lbl.setText("Pending — scan or open tab to check")
            self._gap_status_lbl.setStyleSheet("font-size:11px; color:#585b70;")


# ── Utility ────────────────────────────────────────────────────────────────
def open_path(path: str, folder: bool = False):
    p = Path(path)
    if folder or p.is_dir():
        subprocess.Popen(f'explorer "{p}"', shell=True)
    else:
        subprocess.Popen(f'explorer /select,"{p}"', shell=True)


# ── Main Window ────────────────────────────────────────────────────────────
class MainWindow(QMainWindow):
    def __init__(self, db: Database, user_name: str, user_prefix: str,
                 cat_prefixes: Dict[str, str] = None):
        super().__init__()
        self.db           = db
        self.user_name    = user_name
        self.user_prefix  = user_prefix
        self.cat_prefixes = cat_prefixes or {}
        self._scan_dlg      = None
        self._worker        = None
        self._rescan_worker = None
        self._watcher       = QFileSystemWatcher(self)
        self._watcher.directoryChanged.connect(self._on_dir_changed)
        self._build()
        # Delay scan so window renders first
        QTimer.singleShot(200, self._start_scan)

    def closeEvent(self, event):
        if self._worker and self._worker.isRunning():
            self._worker.terminate()
            self._worker.wait(2000)
        event.accept()
        QApplication.quit()

    def _build(self):
        self.setWindowTitle(f"Parts Tracker — {self.user_name}")
        self.setWindowIcon(make_icon())
        self.setMinimumSize(1300, 700)
        self.resize(1600, 860)

        root = QWidget()
        self.setCentralWidget(root)
        root_ly = QVBoxLayout(root)
        root_ly.setContentsMargins(0, 0, 0, 0)
        root_ly.setSpacing(0)

        # ── Top bar ──
        top = QFrame()
        top.setObjectName("top_bar")
        top.setFixedHeight(54)
        tl = QHBoxLayout(top)
        tl.setContentsMargins(18, 0, 18, 0)
        tl.setSpacing(12)

        app_lbl = QLabel("Parts Tracker")
        app_lbl.setStyleSheet("font-size:16px; font-weight:bold; color:#89b4fa;")
        tl.addWidget(app_lbl)

        sep = QFrame()
        sep.setFrameShape(QFrame.Shape.VLine)
        sep.setStyleSheet("color:#313244;")
        tl.addWidget(sep)

        tl.addStretch()

        self.user_lbl = QLabel(f"Logged in as  {self.user_name}  (prefix: {self.user_prefix})")
        self.user_lbl.setObjectName("lbl_sub")
        tl.addWidget(self.user_lbl)

        self.change_btn = QPushButton("Change User")
        self.change_btn.clicked.connect(self._change_user)
        tl.addWidget(self.change_btn)

        self.clear_btn = QPushButton("Clear All")
        self.clear_btn.setStyleSheet(
            "QPushButton { background:#f38ba8; color:#1e1e2e; font-weight:bold; border:none;"
            " border-radius:6px; padding:6px 14px; }"
            "QPushButton:hover { background:#eba0ac; }"
        )
        self.clear_btn.clicked.connect(self._clear_all)
        tl.addWidget(self.clear_btn)

        self.refresh_btn = QPushButton("  Refresh")
        self.refresh_btn.setObjectName("primary")
        self.refresh_btn.clicked.connect(self._start_scan)
        tl.addWidget(self.refresh_btn)

        root_ly.addWidget(top)

        # ── Tabs ──
        self.tabs = QTabWidget()
        self.tabs.setDocumentMode(True)

        self.tab_my    = MyPartsTab(self.db, self.user_prefix)
        self.tab_all   = AllPartsTab(self.db, self.user_prefix)
        self.tab_jobs  = JobsTab(self.db)
        self.tab_next  = NextNumbersTab(self.db, self.user_prefix, self.cat_prefixes)

        self.tabs.addTab(self.tab_my,   f"  My Parts ({self.user_prefix})  ")
        self.tabs.addTab(self.tab_all,  "  All Parts  ")
        self.tabs.addTab(self.tab_jobs, "  Jobs  ")
        self.tabs.addTab(self.tab_next, "  Next Numbers  ")
        self.tabs.currentChanged.connect(self._tab_changed)

        root_ly.addWidget(self.tabs)

        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage("Ready")

    def _clear_all(self):
        ans = QMessageBox.question(
            self, "Clear All Data",
            "This will delete all scanned jobs and parts from the local database.\n"
            "You can re-scan at any time. Continue?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        if ans == QMessageBox.StandardButton.Yes:
            self.db.clear_all()
            self._reload_tabs()
            self.status_bar.showMessage("Database cleared.", 5000)

    def _start_scan(self):
        if self._worker and self._worker.isRunning():
            return
        self.refresh_btn.setEnabled(False)
        self._scan_dlg = ScanDialog(self)
        self._worker   = ScanWorker(self.db, self.user_prefix)
        self._worker.progress.connect(self._scan_dlg.update)
        self._worker.done.connect(self._scan_done)
        self._worker.err.connect(self._scan_err)
        self._worker.folders_found.connect(self._update_watcher)
        self._worker.start()
        self._scan_dlg.exec()

    def _scan_done(self, new_parts, new_jobs):
        if self._scan_dlg:
            self._scan_dlg.accept()
        self.refresh_btn.setEnabled(True)
        self._reload_tabs()
        self.status_bar.showMessage(
            f"Scan complete — {new_parts:,} part(s) processed, {new_jobs} new job(s)",
            8000
        )

    def _update_watcher(self, folder_paths: list):
        """Replace watched directories with the latest set of scanned folders."""
        old = self._watcher.directories()
        if old:
            self._watcher.removePaths(old)
        if folder_paths:
            self._watcher.addPaths(folder_paths)

    def _on_dir_changed(self, path: str):
        """A watched directory changed — re-scan it in the background."""
        if self._rescan_worker and self._rescan_worker.isRunning():
            return
        self._rescan_worker = DirectoryRescanWorker(self.db, path)
        self._rescan_worker.done.connect(self._rescan_done)
        self._rescan_worker.start()

    def _rescan_done(self):
        self._reload_tabs()
        self.status_bar.showMessage("File change detected — numbers updated.", 4000)

    def _scan_err(self, msg):
        if self._scan_dlg:
            self._scan_dlg.reject()
        self.refresh_btn.setEnabled(True)
        QMessageBox.critical(self, "Scan Error", msg)

    def _tab_changed(self, index: int):
        if self.tabs.widget(index) is self.tab_next:
            self.tab_next.refresh(self.user_prefix, self.cat_prefixes)
            self.tab_next._start_gap_scan()

    def _reload_tabs(self):
        self.tab_my.refresh(self.user_prefix)
        self.tab_all.refresh(self.user_prefix)
        self.tab_jobs.refresh()
        self.tab_next.refresh(self.user_prefix, self.cat_prefixes)
        self.tab_next._start_gap_scan()
        self.tabs.setTabText(0, f"  My Parts ({self.user_prefix})  ")

    def _change_user(self):
        dlg = SetupDialog(self,
                          prefill_id=self.user_prefix if self.user_prefix else "",
                          prefill_cats=self.cat_prefixes)
        if dlg.exec() != QDialog.DialogCode.Accepted:
            return
        v = dlg.values()
        new_prefix     = v["user_prefix"]
        new_cat_pfx    = v["cat_prefixes"]
        prefix_changed = (new_prefix != self.user_prefix)
        self.user_prefix  = new_prefix
        self.cat_prefixes = new_cat_pfx
        self.db.put("user_id",     v["user_id"])
        self.db.put("user_prefix", self.user_prefix)
        self.db.set_cat_prefixes(self.cat_prefixes)
        self.user_lbl.setText(
            f"Logged in as  {self.user_name}  (prefix: {self.user_prefix})"
        )
        if prefix_changed:
            # Main ID changed — clear and re-scan for the new user
            self.db.clear_all()
        self._reload_tabs()
        if prefix_changed:
            self._start_scan()


# ── Entry Point ────────────────────────────────────────────────────────────
def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    app.setStyleSheet(STYLE)
    app.setWindowIcon(make_icon())

    db = Database()

    # Auto-detect Windows username — no need to ask
    user_name   = os.environ.get("USERNAME", "User")
    user_id     = db.get("user_id")
    user_prefix = db.get("user_prefix")

    if not user_id:
        dlg = SetupDialog()
        if dlg.exec() != QDialog.DialogCode.Accepted:
            sys.exit(0)
        v = dlg.values()
        user_id     = v["user_id"]
        user_prefix = v["user_prefix"]
        db.put("user_id",     user_id)
        db.put("user_prefix", user_prefix)
        db.set_cat_prefixes(v["cat_prefixes"])

    if not user_prefix:
        user_prefix = user_id

    cat_prefixes = db.get_cat_prefixes()

    win = MainWindow(db, user_name, user_prefix, cat_prefixes)
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
