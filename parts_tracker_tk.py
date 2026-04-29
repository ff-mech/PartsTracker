"""
Parts Tracker — Tkinter panel embedded in Engineering Tool Hub.

Ported from the PyQt6 Parts Tracker (tools/File Logger/parts_tracker.py) to
run natively inside app.py. Exposes My Parts + Next Numbers tabs only —
matching the ETH2 wrapper scope.

Entry point:
    panel = PartsTrackerPanel(parent, theme)
    panel.pack(fill="both", expand=True)
"""

from __future__ import annotations

import os
import re
import sys
import sqlite3
import threading
import queue
import subprocess
import tkinter as tk
from tkinter import ttk, messagebox
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, List, Tuple, Callable


# ═════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ═════════════════════════════════════════════════════════════════════

JOBS_ROOT       = r"Z:\FOXFAB_DATA\ENGINEERING\2 JOBS"
EVERYTHING_URL  = "http://localhost:8080/"
DB_PATH         = Path(os.environ["APPDATA"]) / "PartsTracker" / "parts.db"

CATEGORIES = {
    "003": "Top Level Assembly",
    "100": "Subassembly",
    "200": "Metal",
    "240": "Copper",
    "245": "Flexibar",
    "250": "Galvanized",
    "295": "Insulation Barrier",
}

# Per-category accent colors (used for card left-borders and Next labels)
CAT_COLORS = {
    "100": "#2563EB",  # blue   – Subassembly
    "200": "#6B7280",  # grey   – Metal
    "240": "#EA580C",  # orange – Copper
    "245": "#16A34A",  # green  – Flexibar
    "250": "#0EA5E9",  # cyan   – Galvanized
    "295": "#D97706",  # amber  – Insulation Barrier
}

PART_RE = re.compile(r"^(\d{3})-(\d{5})((?:_\d+)*)\.(sldprt|sldasm)$", re.IGNORECASE)


def decode_part_filename(name: str) -> Optional["tuple[str, List[str], str]"]:
    """Decode a SolidWorks part filename into category, all covered part numbers,
    and lowercase extension. Combined-part files cover multiple sequential numbers
    via the FoxFab '_NN' suffix convention used in the bom-filler script:
      '240-90129.SLDPRT'         -> ('240', ['240-90129'], 'sldprt')
      '240-90129_30.SLDPRT'      -> ('240', ['240-90129', '240-90130'], 'sldprt')
      '240-90123_124_125.SLDASM' -> ('240', ['240-90123', '240-90124', '240-90125'], 'sldasm')
    Each '_<suffix>' segment replaces the last len(suffix) chars of the previous
    part number. Returns None if the name doesn't match the part-file pattern."""
    m = PART_RE.match(name)
    if not m:
        return None
    cat = m.group(1)
    base = f"{cat}-{m.group(2)}"
    suffix_str = m.group(3) or ""
    ext = m.group(4).lower()
    if not suffix_str:
        return cat, [base], ext
    covered = [base]
    current = base
    for s in (seg for seg in suffix_str.split('_') if seg):
        n = len(s)
        if len(current) < n:
            return cat, [base], ext  # malformed — fall back to base only
        current = current[:-n] + s
        covered.append(current)
    return cat, covered, ext
JOB_RE  = re.compile(r"^(J\d{5})([\s\-].*)?$", re.IGNORECASE)
SUBJ_RE = re.compile(r"^(J\d{5}-\d{2})$",      re.IGNORECASE)

AUTO_RESCAN_MINUTES = 5  # SMB watcher fallback — ticks a silent background scan
DB_POLL_MS          = 5000   # Next Numbers tab refresh cadence


# ═════════════════════════════════════════════════════════════════════
#  DATABASE
# ═════════════════════════════════════════════════════════════════════

class Database:
    def __init__(self):
        DB_PATH.parent.mkdir(parents=True, exist_ok=True)
        self.con = sqlite3.connect(str(DB_PATH), check_same_thread=False)
        self.con.row_factory = sqlite3.Row
        self._init()

    def _init(self):
        self.con.executescript("""
            CREATE TABLE IF NOT EXISTS settings (key TEXT PRIMARY KEY, value TEXT);

            CREATE TABLE IF NOT EXISTS jobs (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                job_number      TEXT NOT NULL,
                job_name        TEXT,
                sub_job         TEXT NOT NULL,
                catalog_no      TEXT,
                enclosure_size  TEXT,
                prf_path        TEXT,
                scanned_at      TEXT,
                is_archived     INTEGER DEFAULT 0,
                UNIQUE(job_number, sub_job)
            );

            CREATE TABLE IF NOT EXISTS parts (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                part_number   TEXT NOT NULL,
                category_code TEXT,
                category_name TEXT,
                user_prefix   TEXT,
                file_ext      TEXT,
                full_path     TEXT NOT NULL UNIQUE,
                job_id        INTEGER REFERENCES jobs(id)
            );

            CREATE INDEX IF NOT EXISTS idx_p_user ON parts(user_prefix);
            CREATE INDEX IF NOT EXISTS idx_p_job  ON parts(job_id);
            CREATE INDEX IF NOT EXISTS idx_p_cat  ON parts(category_code);

            CREATE TABLE IF NOT EXISTS part_history (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                event       TEXT NOT NULL,
                part_number TEXT NOT NULL,
                full_path   TEXT NOT NULL,
                job_id      INTEGER,
                ts          TEXT NOT NULL
            );
            CREATE INDEX IF NOT EXISTS idx_h_ts ON part_history(ts);
            CREATE INDEX IF NOT EXISTS idx_h_pn ON part_history(part_number);
        """)
        self.con.commit()
        try:
            self.con.execute("ALTER TABLE jobs ADD COLUMN is_archived INTEGER DEFAULT 0")
            self.con.commit()
        except Exception:
            pass

        # Migration: drop UNIQUE on parts.full_path so combined-part files
        # (e.g. 240-90129_30.sldprt) can occupy multiple part numbers via
        # one row per covered number with shared full_path.
        self._migrate_parts_uniqueness()

    def _migrate_parts_uniqueness(self):
        cur = self.con.cursor()
        for idx in cur.execute("PRAGMA index_list(parts)").fetchall():
            if not idx["unique"]:
                continue
            cols = [c["name"] for c in cur.execute(f"PRAGMA index_info({idx['name']})").fetchall()]
            if cols != ["full_path"]:
                continue
            cur.executescript("""
                CREATE TABLE parts_new (
                    id            INTEGER PRIMARY KEY AUTOINCREMENT,
                    part_number   TEXT NOT NULL,
                    category_code TEXT,
                    category_name TEXT,
                    user_prefix   TEXT,
                    file_ext      TEXT,
                    full_path     TEXT NOT NULL,
                    job_id        INTEGER REFERENCES jobs(id),
                    UNIQUE(full_path, part_number)
                );
                INSERT INTO parts_new (id, part_number, category_code, category_name,
                                        user_prefix, file_ext, full_path, job_id)
                    SELECT id, part_number, category_code, category_name,
                           user_prefix, file_ext, full_path, job_id FROM parts;
                DROP TABLE parts;
                ALTER TABLE parts_new RENAME TO parts;
                CREATE INDEX IF NOT EXISTS idx_p_user ON parts(user_prefix);
                CREATE INDEX IF NOT EXISTS idx_p_job  ON parts(job_id);
                CREATE INDEX IF NOT EXISTS idx_p_cat  ON parts(category_code);
            """)
            self.con.commit()
            return

    # ── settings ──
    def get(self, key, default=None):
        r = self.con.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone()
        return r["value"] if r else default

    def put(self, key, value):
        self.con.execute("INSERT OR REPLACE INTO settings VALUES(?,?)", (key, str(value)))
        self.con.commit()

    # ── jobs ──
    def job_id(self, job_number, sub_job):
        r = self.con.execute(
            "SELECT id FROM jobs WHERE job_number=? AND sub_job=?", (job_number, sub_job)
        ).fetchone()
        return r["id"] if r else None

    def upsert_job(self, job_number, job_name, sub_job,
                   catalog_no=None, enclosure_size=None, prf_path=None, is_archived=0):
        self.con.execute("""
            INSERT INTO jobs(job_number,job_name,sub_job,catalog_no,enclosure_size,prf_path,scanned_at,is_archived)
            VALUES(?,?,?,?,?,?,?,?)
            ON CONFLICT(job_number,sub_job) DO UPDATE SET
                job_name       = excluded.job_name,
                catalog_no     = COALESCE(excluded.catalog_no,    catalog_no),
                enclosure_size = COALESCE(excluded.enclosure_size, enclosure_size),
                prf_path       = COALESCE(excluded.prf_path,      prf_path),
                scanned_at     = excluded.scanned_at,
                is_archived    = excluded.is_archived
        """, (job_number, job_name, sub_job, catalog_no, enclosure_size, prf_path,
              datetime.now().isoformat(), is_archived))
        self.con.commit()
        return self.job_id(job_number, sub_job)

    def upsert_part(self, part_numbers, category_code, category_name,
                    file_ext, full_path, job_id):
        """Insert/update one row per covered part number for a single file.
        `part_numbers` may be a single 'CAT-NNNNN' string or a list of them
        (combined-part files cover multiple sequential numbers)."""
        if isinstance(part_numbers, str):
            part_numbers = [part_numbers]
        ts = datetime.now().isoformat()
        for pn in part_numbers:
            five = pn.split('-', 1)[1] if '-' in pn else pn
            is_new = not self.con.execute(
                "SELECT 1 FROM parts WHERE full_path=? AND part_number=?",
                (full_path, pn)
            ).fetchone()
            self.con.execute("""
                INSERT INTO parts(part_number,category_code,category_name,
                                  user_prefix,file_ext,full_path,job_id)
                VALUES(?,?,?,?,?,?,?)
                ON CONFLICT(full_path, part_number) DO UPDATE SET job_id=excluded.job_id
            """, (pn, category_code, category_name, five, file_ext, full_path, job_id))
            if is_new:
                self.con.execute(
                    "INSERT INTO part_history(event,part_number,full_path,job_id,ts)"
                    " VALUES(?,?,?,?,?)",
                    ("add", pn, full_path, job_id, ts)
                )
        self.con.commit()

    def get_jobs(self, search=""):
        q = """SELECT j.*, COUNT(p.id) as part_count
               FROM jobs j LEFT JOIN parts p ON p.job_id=j.id WHERE 1=1"""
        args = []
        if search:
            q += " AND (j.job_number LIKE ? OR j.job_name LIKE ? OR j.sub_job LIKE ?)"
            args += [f"%{search}%"] * 3
        q += " GROUP BY j.id ORDER BY j.job_number, j.sub_job"
        return self.con.execute(q, args).fetchall()

    def get_parts(self, category="", job_id=None, file_ext=""):
        q = """SELECT p.*, j.job_number, j.job_name, j.sub_job,
                      j.catalog_no, j.enclosure_size, j.is_archived
               FROM parts p LEFT JOIN jobs j ON p.job_id=j.id WHERE 1=1"""
        args = []
        if category:
            q += " AND p.category_code=?"
            args.append(category)
        if job_id is not None:
            q += " AND p.job_id=?"
            args.append(job_id)
        if file_ext:
            q += " AND p.file_ext=?"
            args.append(file_ext)
        q += " ORDER BY p.part_number"
        return self.con.execute(q, args).fetchall()

    def get_cat_prefixes(self) -> Dict[str, str]:
        result = {}
        for code in CATEGORIES:
            if code == "003":
                continue
            val = self.get(f"cat_prefix_{code}")
            if val:
                result[code] = val
        return result

    def set_cat_prefixes(self, cat_prefixes: Dict[str, str]):
        for code in CATEGORIES:
            if code == "003":
                continue
            val = cat_prefixes.get(code, "")
            if val:
                self.put(f"cat_prefix_{code}", val)
            else:
                self.con.execute("DELETE FROM settings WHERE key=?", (f"cat_prefix_{code}",))
        self.con.commit()

    @staticmethod
    def _prefix_range(user_prefix: str):
        p = int(user_prefix)
        pad = 5 - len(user_prefix)
        lo = p * (10 ** pad)
        hi = (p + 1) * (10 ** pad) - 1
        return lo, hi

    def latest_by_category(self, user_prefix: str = "",
                            cat_prefixes: Dict[str, str] = None) -> List[Dict]:
        if cat_prefixes is None:
            cat_prefixes = {}
        con = sqlite3.connect(str(DB_PATH), timeout=10)
        con.row_factory = sqlite3.Row

        result: Dict[str, dict] = {}
        for cat_code, cat_name in CATEGORIES.items():
            if cat_code == "003":
                continue
            pfx = cat_prefixes.get(cat_code, user_prefix)
            if pfx:
                lo, hi = self._prefix_range(pfx)
                rows = con.execute("""
                    SELECT part_number, full_path
                    FROM parts
                    WHERE category_code = ?
                      AND CAST(user_prefix AS INTEGER) BETWEEN ? AND ?
                    ORDER BY part_number DESC
                """, (cat_code, lo, hi)).fetchall()
            else:
                rows = con.execute("""
                    SELECT part_number, full_path
                    FROM parts
                    WHERE category_code = ?
                    ORDER BY part_number DESC
                """, (cat_code,)).fetchall()

            for row in rows:
                if Path(row["full_path"]).exists():
                    result[cat_code] = {
                        "category_code": cat_code,
                        "category_name": cat_name,
                        "latest_part":   row["part_number"],
                        "effective_pfx": pfx,
                    }
                    break

        con.close()
        return list(result.values())

    def get_duplicate_parts(self, user_prefix: str) -> List[dict]:
        if not user_prefix:
            return []
        try:
            lo, hi = self._prefix_range(user_prefix)
        except ValueError:
            return []
        NO_ARCHIVE = "AND LOWER(full_path) NOT LIKE '%\\archive\\%'"
        rows = self.con.execute(
            f"""
            SELECT p.part_number, p.file_ext, p.full_path,
                   j.job_number, j.sub_job, j.job_name
            FROM parts p
            LEFT JOIN jobs j ON p.job_id = j.id
            INNER JOIN (
                SELECT part_number, file_ext
                FROM parts
                WHERE CAST(user_prefix AS INTEGER) BETWEEN ? AND ?
                  {NO_ARCHIVE}
                GROUP BY part_number, file_ext
                HAVING COUNT(DISTINCT full_path) > 1
            ) dup
              ON dup.part_number = p.part_number
             AND dup.file_ext    = p.file_ext
            WHERE CAST(p.user_prefix AS INTEGER) BETWEEN ? AND ?
              {NO_ARCHIVE.replace('full_path', 'p.full_path')}
            ORDER BY p.part_number, p.file_ext, p.full_path
            """,
            (lo, hi, lo, hi),
        ).fetchall()

        groups: Dict[Tuple[str, str], dict] = {}
        for r in rows:
            key = (r["part_number"], r["file_ext"])
            g = groups.setdefault(key, {
                "part_number": r["part_number"],
                "file_ext":    r["file_ext"],
                "entries":     [],
            })
            if any(e["full_path"] == r["full_path"] for e in g["entries"]):
                continue
            g["entries"].append({
                "full_path":  r["full_path"],
                "job_number": r["job_number"],
                "sub_job":    r["sub_job"],
                "job_name":   r["job_name"],
            })
        return [g for g in groups.values() if len(g["entries"]) > 1]

    def clear_all(self):
        self.con.executescript("DELETE FROM parts; DELETE FROM jobs;")
        self.con.commit()

    def close(self):
        try:
            self.con.close()
        except Exception:
            pass


# ═════════════════════════════════════════════════════════════════════
#  EVERYTHING / PRF / PARSING HELPERS
# ═════════════════════════════════════════════════════════════════════

def _eq(session, query: str) -> List[Dict]:
    """One Everything HTTP query."""
    r = session.get(EVERYTHING_URL, params={
        "s": query, "j": 1, "path_column": 1, "count": 200000,
    }, timeout=60)
    r.raise_for_status()
    out = []
    for item in r.json().get("results", []):
        name = item.get("name") or item.get("filename", "")
        path = item.get("path", "")
        if name and path:
            out.append({"name": name, "path": path})
    return out


def find_003_folders(user_prefix: str) -> List[Path]:
    import requests
    session = requests.Session()
    try:
        hits = _eq(session, f'"003-{user_prefix}" ext:sldasm path:"{JOBS_ROOT}"')
        if not hits:
            raise RuntimeError(
                f"No '003-{user_prefix}*.sldasm' files found.\n\n"
                f"Searched in: {JOBS_ROOT}\n\n"
                "Check that:\n"
                "1. Everything HTTP Server is enabled\n"
                "   (Everything → Tools → Options → HTTP Server, port 8080)\n"
                "2. Everything has indexed the Z:\\ drive\n"
                f"3. Your user prefix is correct (current: {user_prefix})"
            )
        return list({Path(h["path"]) for h in hits})
    except requests.exceptions.ConnectionError:
        raise ConnectionError(
            "Cannot connect to Everything HTTP server.\n\n"
            "Enable it in Everything:\n"
            "Tools → Options → HTTP Server → Enable (port 8080)"
        )


def find_gaps_via_everything(user_prefix: str,
                              cat_prefixes: Dict[str, str] = None) -> Dict[str, List[str]]:
    """Compute real gaps by asking Everything for every matching file on disk."""
    if not user_prefix:
        return {}
    if cat_prefixes is None:
        cat_prefixes = {}
    try:
        import requests
        session = requests.Session()
        result: Dict[str, List[str]] = {}
        for cat_code in CATEGORIES:
            if cat_code == "003":
                continue
            pfx = cat_prefixes.get(cat_code, user_prefix)
            lo, hi = Database._prefix_range(pfx)

            hits: List[Dict] = []
            for ext in ("sldprt", "sldasm"):
                hits += _eq(
                    session,
                    f'"{cat_code}-{pfx}" ext:{ext} path:"{JOBS_ROOT}"',
                )
            # Combined-part files (e.g. 240-90129_30) contribute every covered number.
            present: set = set()
            for hit in hits:
                decoded = decode_part_filename(hit["name"])
                if not decoded or decoded[0] != cat_code:
                    continue
                for pn in decoded[1]:
                    try:
                        n = int(pn.split('-', 1)[1])
                    except (ValueError, IndexError):
                        continue
                    if lo <= n <= hi:
                        present.add(n)

            if len(present) < 2:
                continue
            min_n, max_n = min(present), max(present)
            gaps = [
                f"{cat_code}-{str(n).zfill(5)}"
                for n in range(min_n + 1, max_n)
                if n not in present
            ]
            if gaps:
                result[cat_code] = gaps
        return result
    except Exception:
        return {}


def is_part_number_taken(part_text: str, timeout: float = 3.0) -> Optional[str]:
    """Click-time check: ask Everything whether `part_text` (e.g. "240-90130")
    is occupied by any .sldprt/.sldasm file under JOBS_ROOT, INCLUDING combined-part
    files like "240-90129_30.sldprt" that cover the requested number.
    Returns the conflicting full path, or None if free / on Everything error.
    Fail-open on error so a flaky Everything doesn't block the user."""
    m = re.match(r'^(\d{3})-(\d{5})$', part_text)
    if not m:
        return None
    cat_code = m.group(1)
    five_digit = m.group(2)
    # Search the wider "CAT-D" block (D = first digit of the 5-digit) so we catch
    # combined siblings whose filenames don't contain the literal target string.
    block_prefix = five_digit[0]
    try:
        import requests
        for ext in ("sldprt", "sldasm"):
            r = requests.get(EVERYTHING_URL, params={
                "s": f'"{cat_code}-{block_prefix}" ext:{ext} path:"{JOBS_ROOT}"',
                "j": 1, "path_column": 1, "count": 200000,
            }, timeout=timeout)
            r.raise_for_status()
            for item in r.json().get("results", []):
                name = item.get("name") or item.get("filename", "")
                path = item.get("path", "")
                if not (name and path):
                    continue
                decoded = decode_part_filename(name)
                if not decoded or decoded[0] != cat_code:
                    continue
                if part_text in decoded[1]:
                    return str(Path(path) / name)
        return None
    except Exception:
        return None


def find_prf(job_root: Path, sub_job: str) -> Optional[str]:
    prf_dir = job_root / "300 Inputs" / "302 Production Release Form"
    if not prf_dir.exists():
        return None
    sub_lower = sub_job.lower()
    candidates = [f for f in prf_dir.iterdir()
                  if f.suffix.lower() == ".xlsx" and "prf" in f.stem.lower()]
    for f in candidates:
        if sub_lower in f.stem.lower():
            return str(f)
    return str(candidates[0]) if candidates else None


def read_prf(path: str) -> Dict:
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        catalog = str(ws["C6"].value or "").strip() or None
        size_raw = str(ws["G18"].value or "").strip()
        m = re.search(r"(\d+)\s*[xX×]\s*(\d+)\s*[xX×]\s*(\d+)", size_raw)
        size = f"{m.group(1)} x {m.group(2)} x {m.group(3)}" if m else (size_raw or None)
        wb.close()
        return {"catalog_no": catalog, "enclosure_size": size}
    except Exception:
        return {"catalog_no": None, "enclosure_size": None}


def _next_part(user_prefix: str, cat_code: str, latest: str) -> str:
    if latest:
        five = latest.split("-")[1]
        nxt  = str(int(five) + 1).zfill(5)
    else:
        zeros = "0" * (4 - len(user_prefix))
        nxt   = f"{user_prefix}{zeros}1"
    return f"{cat_code}-{nxt}"


def open_path(path: str, folder: bool = False):
    p = Path(path)
    try:
        if folder or p.is_dir():
            subprocess.Popen(f'explorer "{p}"', shell=True)
        else:
            subprocess.Popen(f'explorer /select,"{p}"', shell=True)
    except Exception:
        pass


# ═════════════════════════════════════════════════════════════════════
#  WORKER THREADS (threading + queue-based messaging)
# ═════════════════════════════════════════════════════════════════════
#
# Messages put on the queue are 2-tuples: (tag, payload). Consumers drain
# the queue on the Tk main thread via root.after() polling.
#
# Tags used:
#   ("scan:progress", (pct:int, msg:str))
#   ("scan:done",     (new_parts:int, new_jobs:int))
#   ("scan:err",      msg:str)
#   ("scan:folders",  [str,...])
#   ("scan:cancel",   None)
#   ("gaps:done",     {cat_code: [gap,...]})


class ScanWorker(threading.Thread):
    def __init__(self, db: Database, user_prefix: str, q: queue.Queue):
        super().__init__(daemon=True)
        self.db = db
        self.user_prefix = user_prefix
        self.q = q
        self._cancel = False

    def cancel(self):
        self._cancel = True

    def run(self):
        try:
            self.q.put(("scan:progress", (2, f"Searching for 003-{self.user_prefix}* assemblies…")))
            folders = find_003_folders(self.user_prefix)
            if self._cancel:
                self.q.put(("scan:cancel", None))
                return
            total = len(folders)
            self.q.put(("scan:progress", (5, f"Found {total} part folder(s) — processing…")))

            new_parts = new_jobs = 0

            for i, folder in enumerate(folders):
                if self._cancel:
                    self.q.put(("scan:cancel", None))
                    return
                pct = 5 + int(90 * i / max(1, total))

                parts_list = list(folder.parts)
                job_number = job_name_str = sub_job = None
                job_root_path = None
                for idx, part in enumerate(parts_list):
                    if SUBJ_RE.match(part):
                        sub_job = part.upper()
                    elif JOB_RE.match(part) and not SUBJ_RE.match(part):
                        jm = JOB_RE.match(part)
                        job_number = jm.group(1).upper()
                        rest = (jm.group(2) or "").strip(" -").strip()
                        job_name_str = rest or None
                        job_root_path = Path(*parts_list[: idx + 1])
                if not job_number:
                    continue
                if not sub_job:
                    sub_job = job_number

                is_archived = any(p.lower() == "archive" for p in parts_list)

                existing_id = self.db.job_id(job_number, sub_job)
                if existing_id:
                    job_id = existing_id
                    self.db.con.execute(
                        "UPDATE jobs SET is_archived=? WHERE id=?",
                        (1 if is_archived else 0, existing_id)
                    )
                    self.db.con.commit()
                else:
                    self.q.put(("scan:progress", (pct, f"New job {job_number} — reading PRF…")))
                    prf_data = {"catalog_no": None, "enclosure_size": None}
                    prf_path = None
                    if job_root_path and job_root_path.exists():
                        prf_path = find_prf(job_root_path, sub_job)
                        if prf_path:
                            prf_data = read_prf(prf_path)
                    job_id = self.db.upsert_job(
                        job_number, job_name_str, sub_job,
                        prf_data["catalog_no"], prf_data["enclosure_size"],
                        prf_path, is_archived=1 if is_archived else 0
                    )
                    new_jobs += 1

                self.q.put(("scan:progress", (pct, f"Scanning {job_number}/{sub_job}…")))
                try:
                    files = [f for f in folder.iterdir()
                             if f.is_file() and f.suffix.lower() in (".sldprt", ".sldasm")]
                except Exception:
                    continue

                for f in files:
                    decoded = decode_part_filename(f.name)
                    if not decoded:
                        continue
                    cat_code, part_nums, file_ext = decoded
                    cat_name = CATEGORIES.get(cat_code, "Unknown")
                    self.db.upsert_part(
                        part_nums, cat_code, cat_name,
                        file_ext, str(f), job_id
                    )
                    new_parts += len(part_nums)

            self.q.put(("scan:progress", (100, "Complete")))
            self.q.put(("scan:folders", [str(f) for f in folders]))
            self.q.put(("scan:done", (new_parts, new_jobs)))

        except Exception as e:
            if self._cancel:
                self.q.put(("scan:cancel", None))
            else:
                self.q.put(("scan:err", str(e)))


class GapScanWorker(threading.Thread):
    def __init__(self, user_prefix: str, cat_prefixes: Dict[str, str], q: queue.Queue):
        super().__init__(daemon=True)
        self.user_prefix = user_prefix
        self.cat_prefixes = cat_prefixes
        self.q = q

    def run(self):
        result = find_gaps_via_everything(self.user_prefix, self.cat_prefixes)
        self.q.put(("gaps:done", result))


# ═════════════════════════════════════════════════════════════════════
#  DIALOGS (Toplevel)
# ═════════════════════════════════════════════════════════════════════

class ScanDialog(tk.Toplevel):
    """Modal progress dialog shown during a user-triggered scan."""

    def __init__(self, parent, theme: dict, on_cancel: Callable[[], None]):
        super().__init__(parent)
        self.theme = theme
        self._on_cancel = on_cancel

        self.title("Scanning…")
        self.configure(bg=theme["BG"])
        self.resizable(False, False)
        self.transient(parent.winfo_toplevel())
        try:
            self.grab_set()
        except tk.TclError:
            pass

        w, h = 480, 170
        self.geometry(f"{w}x{h}")
        self.update_idletasks()
        try:
            px = parent.winfo_rootx() + (parent.winfo_width() - w) // 2
            py = parent.winfo_rooty() + (parent.winfo_height() - h) // 2
            self.geometry(f"{w}x{h}+{max(0, px)}+{max(0, py)}")
        except Exception:
            pass

        outer = tk.Frame(self, bg=theme["BG"])
        outer.pack(fill="both", expand=True, padx=24, pady=20)

        self.lbl = tk.Label(outer, text="Initializing…", bg=theme["BG"],
                            fg=theme["SUBTLE"], font=theme["BODY"], anchor="w")
        self.lbl.pack(fill="x")

        self.bar_var = tk.IntVar(value=0)
        self.bar = ttk.Progressbar(outer, orient="horizontal", mode="determinate",
                                    maximum=100, variable=self.bar_var,
                                    style="PT.Horizontal.TProgressbar")
        self.bar.pack(fill="x", pady=(10, 6))

        self.detail = tk.Label(outer, text="", bg=theme["BG"],
                               fg=theme["SUBTLE"], font=theme["SMALL"], anchor="w")
        self.detail.pack(fill="x")

        btn_row = tk.Frame(outer, bg=theme["BG"])
        btn_row.pack(fill="x", pady=(8, 0))
        self.cancel_btn = tk.Button(
            btn_row, text="Cancel",
            bg=theme["BG"], fg=theme["TEXT"], font=theme["BODY"],
            relief="flat", cursor="hand2",
            activebackground=theme["BORDER"],
            command=self._cancel,
        )
        self.cancel_btn.pack(side="right")

        self.protocol("WM_DELETE_WINDOW", self._cancel)

    def _cancel(self):
        self.cancel_btn.configure(state="disabled")
        self.lbl.configure(text="Cancelling…")
        try:
            self._on_cancel()
        except Exception:
            pass

    def update_progress(self, pct: int, msg: str):
        try:
            self.bar_var.set(max(0, min(100, int(pct))))
            self.lbl.configure(text=msg)
        except tk.TclError:
            pass


class SetupDialog(tk.Toplevel):
    """Modal setup form — user ID + optional per-category overrides."""

    def __init__(self, parent, theme: dict,
                 prefill_id: str = "", prefill_cats: Dict[str, str] = None):
        super().__init__(parent)
        self.theme = theme
        self._result: Optional[dict] = None
        self._prefill_cats = prefill_cats or {}
        self._cat_vars: Dict[str, tk.StringVar] = {}
        self._more_open = False

        self.title("Parts Tracker — Setup")
        self.configure(bg=theme["BG"])
        self.resizable(False, False)
        self.transient(parent.winfo_toplevel())
        try:
            self.grab_set()
        except tk.TclError:
            pass

        outer = tk.Frame(self, bg=theme["BG"])
        outer.pack(fill="both", expand=True, padx=28, pady=22)

        tk.Label(outer, text="User Setup",
                 bg=theme["BG"], fg=theme["ACCENT"],
                 font=("Segoe UI", 14, "bold")).pack(pady=(0, 4))
        tk.Label(outer, text="Enter your SolidWorks user ID.",
                 bg=theme["BG"], fg=theme["SUBTLE"],
                 font=theme["SMALL"]).pack(pady=(0, 14))

        # Main ID form
        form = tk.Frame(outer, bg=theme["BG"])
        form.pack(fill="x")

        tk.Label(form, text="User ID:", bg=theme["BG"], fg=theme["TEXT"],
                 font=theme["BODY"], width=14, anchor="e").grid(row=0, column=0, sticky="e", padx=(0, 8), pady=4)
        self.id_var = tk.StringVar(value=prefill_id)
        self.id_entry = tk.Entry(form, textvariable=self.id_var,
                                 font=theme["BODY"],
                                 bg="#F8FAFC", fg=theme["TEXT"],
                                 relief="flat", highlightthickness=1,
                                 highlightbackground=theme["BORDER"],
                                 highlightcolor=theme["ACCENT"])
        self.id_entry.grid(row=0, column=1, sticky="ew", pady=4)
        form.columnconfigure(1, weight=1)

        tk.Label(form, text="Search prefix:", bg=theme["BG"], fg=theme["TEXT"],
                 font=theme["BODY"], width=14, anchor="e").grid(row=1, column=0, sticky="e", padx=(0, 8), pady=4)
        self.preview = tk.Label(form, text="", bg=theme["BG"],
                                fg=theme["SUCCESS"], font=theme["BODY"], anchor="w")
        self.preview.grid(row=1, column=1, sticky="w", pady=4)
        self.id_var.trace_add("write", lambda *_: self._update_preview())
        self._update_preview()

        # More Options toggle
        self._more_btn = tk.Button(
            outer, text="▶   More Options  —  per-category ID overrides",
            bg=theme["BG"], fg=theme["SUBTLE"], font=theme["SMALL"],
            relief="flat", cursor="hand2", anchor="w",
            activebackground=theme["BG"],
            command=self._toggle_more,
        )
        self._more_btn.pack(fill="x", pady=(14, 4))

        # Collapsible section
        self._more_frame = tk.Frame(outer, bg=theme["BG"])

        tk.Label(self._more_frame,
                 text=("If you were assigned a different ID for specific categories,"
                       " enter it here.\nLeave blank to use the main User ID above."),
                 bg=theme["BG"], fg=theme["SUBTLE"],
                 font=theme["SMALL"], justify="left").pack(anchor="w", pady=(0, 6))

        cat_grid = tk.Frame(self._more_frame, bg=theme["BG"])
        cat_grid.pack(fill="x")
        row = 0
        for code, name in CATEGORIES.items():
            if code == "003":
                continue
            tk.Label(cat_grid, text=f"{code}  {name}:",
                     bg=theme["BG"], fg=theme["TEXT"],
                     font=theme["BODY"], anchor="e", width=22).grid(row=row, column=0, sticky="e", padx=(0, 8), pady=2)
            var = tk.StringVar(value=self._prefill_cats.get(code, ""))
            self._cat_vars[code] = var
            ent = tk.Entry(cat_grid, textvariable=var, font=theme["BODY"],
                           bg="#F8FAFC", fg=theme["TEXT"],
                           relief="flat", highlightthickness=1,
                           highlightbackground=theme["BORDER"],
                           highlightcolor=theme["ACCENT"])
            ent.grid(row=row, column=1, sticky="ew", pady=2)
            cat_grid.columnconfigure(1, weight=1)
            row += 1

        if any(self._prefill_cats.get(c) for c in self._cat_vars):
            self._toggle_more()

        # Save button
        btn_row = tk.Frame(outer, bg=theme["BG"])
        btn_row.pack(fill="x", pady=(14, 0))
        tk.Button(
            btn_row, text="  Save & Continue  ",
            bg=theme["ACCENT"], fg="white",
            activebackground=theme["ACCENT_H"], activeforeground="white",
            font=("Segoe UI", 10, "bold"), relief="flat", cursor="hand2",
            command=self._ok,
        ).pack(side="right")

        self.update_idletasks()
        self.minsize(460, self.winfo_reqheight())
        # center over parent
        try:
            top = parent.winfo_toplevel()
            px = top.winfo_rootx() + (top.winfo_width() - self.winfo_width()) // 2
            py = top.winfo_rooty() + (top.winfo_height() - self.winfo_height()) // 2
            self.geometry(f"+{max(0, px)}+{max(0, py)}")
        except Exception:
            pass

        self.id_entry.focus_set()
        self.bind("<Return>", lambda _e: self._ok())
        self.bind("<Escape>", lambda _e: self._cancel())
        self.protocol("WM_DELETE_WINDOW", self._cancel)

    def _toggle_more(self):
        self._more_open = not self._more_open
        if self._more_open:
            self._more_btn.configure(text="▼   More Options  —  per-category ID overrides")
            self._more_frame.pack(fill="x", pady=(0, 4))
        else:
            self._more_btn.configure(text="▶   More Options  —  per-category ID overrides")
            self._more_frame.pack_forget()
        self.update_idletasks()

    def _update_preview(self):
        uid = self.id_var.get().strip()
        pad = "X" * max(0, 5 - len(uid))
        if uid.isdigit():
            self.preview.configure(text=f"###-{uid}{pad}   (e.g. 003-{uid}{pad})")
        else:
            self.preview.configure(text="")

    def _ok(self):
        uid = self.id_var.get().strip()
        if not uid.isdigit():
            messagebox.showwarning("Invalid", "User ID must be a number.", parent=self)
            return
        cat_prefixes: Dict[str, str] = {}
        for code, var in self._cat_vars.items():
            v = var.get().strip()
            if v:
                if not v.isdigit():
                    messagebox.showwarning(
                        "Invalid",
                        f"Override for category {code} must be a number (or leave blank).",
                        parent=self,
                    )
                    return
                cat_prefixes[code] = v
        self._result = {"user_id": uid, "user_prefix": uid, "cat_prefixes": cat_prefixes}
        self.destroy()

    def _cancel(self):
        self._result = None
        self.destroy()

    @property
    def result(self) -> Optional[dict]:
        return self._result


# ═════════════════════════════════════════════════════════════════════
#  MY PARTS TAB
# ═════════════════════════════════════════════════════════════════════

class MyPartsTab(tk.Frame):
    """Two-pane view: jobs (left) and parts for selected job (right)."""

    PART_COLS = ["part_number", "category", "type", "job", "sub_job",
                 "catalog", "size"]

    def __init__(self, parent, db: Database, user_prefix: str, theme: dict):
        super().__init__(parent, bg=theme["BG"])
        self.db = db
        self.user_prefix = user_prefix
        self.theme = theme
        self._cur_job_id: Optional[int] = None
        self._job_index: Dict[str, dict] = {}   # treeview iid → job row dict
        self._build()
        self._load_jobs()

    # ── layout ──
    def _build(self):
        t = self.theme
        # Horizontal paned window lets users resize the sidebar
        paned = ttk.Panedwindow(self, orient="horizontal", style="PT.TPanedwindow")
        paned.pack(fill="both", expand=True, padx=10, pady=10)

        # ── Left: Jobs sidebar ──
        left = tk.Frame(paned, bg=t["BG"])
        paned.add(left, weight=0)

        tk.Label(left, text="Jobs", bg=t["BG"], fg=t["ACCENT"],
                 font=("Segoe UI", 12, "bold")).pack(anchor="w")

        self.job_search_var = tk.StringVar()
        search = tk.Entry(left, textvariable=self.job_search_var,
                          font=t["BODY"], bg="#F8FAFC", fg=t["TEXT"],
                          relief="flat", highlightthickness=1,
                          highlightbackground=t["BORDER"],
                          highlightcolor=t["ACCENT"])
        search.pack(fill="x", pady=(4, 6))
        search.insert(0, "Filter jobs…")
        search.configure(fg=t["SUBTLE"])

        def _on_focus_in(_e):
            if search.get() == "Filter jobs…":
                search.delete(0, "end")
                search.configure(fg=t["TEXT"])

        def _on_focus_out(_e):
            if not search.get().strip():
                search.configure(fg=t["SUBTLE"])
                search.insert(0, "Filter jobs…")

        search.bind("<FocusIn>", _on_focus_in)
        search.bind("<FocusOut>", _on_focus_out)
        self.job_search_var.trace_add("write", lambda *_: self._on_search_changed())

        # Treeview for jobs (single column, scrollable)
        job_frame = tk.Frame(left, bg=t["BG"], highlightthickness=1,
                             highlightbackground=t["BORDER"])
        job_frame.pack(fill="both", expand=True)

        self.job_tree = ttk.Treeview(job_frame, columns=("job",), show="tree",
                                     style="PT.Treeview", selectmode="browse")
        self.job_tree.heading("#0", text="")
        self.job_tree.column("#0", width=260, stretch=True)
        jy = ttk.Scrollbar(job_frame, orient="vertical", command=self.job_tree.yview)
        self.job_tree.configure(yscrollcommand=jy.set)
        self.job_tree.pack(side="left", fill="both", expand=True)
        jy.pack(side="right", fill="y")
        self.job_tree.bind("<<TreeviewSelect>>", self._on_job_select)

        self.job_count_lbl = tk.Label(left, text="", bg=t["BG"],
                                      fg=t["SUBTLE"], font=t["SMALL"], anchor="w")
        self.job_count_lbl.pack(fill="x", pady=(6, 0))

        # ── Right: Parts panel ──
        right = tk.Frame(paned, bg=t["BG"])
        paned.add(right, weight=1)

        # Header row
        hdr = tk.Frame(right, bg=t["BG"])
        hdr.pack(fill="x")
        self.parts_title_var = tk.StringVar(value="Select a job")
        tk.Label(hdr, textvariable=self.parts_title_var,
                 bg=t["BG"], fg=t["ACCENT"],
                 font=("Segoe UI", 12, "bold"), anchor="w").pack(side="left")

        filt = tk.Frame(hdr, bg=t["BG"])
        filt.pack(side="right")

        cat_items = ["All Categories"] + [f"{c} – {n}" for c, n in CATEGORIES.items()]
        self.cat_var = tk.StringVar(value=cat_items[0])
        self.cat_cb = ttk.Combobox(filt, textvariable=self.cat_var, values=cat_items,
                                   state="readonly", width=24, style="PT.TCombobox")
        self.cat_cb.pack(side="left", padx=(0, 6))
        self.cat_cb.bind("<<ComboboxSelected>>", lambda _e: self._load_parts())

        self.type_var = tk.StringVar(value="All Types")
        self.type_cb = ttk.Combobox(filt, textvariable=self.type_var,
                                    values=["All Types", "sldprt", "sldasm"],
                                    state="readonly", width=12, style="PT.TCombobox")
        self.type_cb.pack(side="left")
        self.type_cb.bind("<<ComboboxSelected>>", lambda _e: self._load_parts())

        # Info strip
        info = tk.Frame(right, bg=t["PANEL"], highlightthickness=1,
                        highlightbackground=t["BORDER"])
        info.pack(fill="x", pady=(6, 6))
        self.info_var = tk.StringVar(value="")
        self.info_lbl = tk.Label(info, textvariable=self.info_var, bg=t["PANEL"],
                                 fg=t["SUBTLE"], font=t["SMALL"],
                                 anchor="w", padx=12, pady=6)
        self.info_lbl.pack(side="left")
        self.part_count_lbl = tk.Label(info, text="", bg=t["PANEL"],
                                       fg=t["SUCCESS"], font=t["SMALL"],
                                       anchor="e", padx=12)
        self.part_count_lbl.pack(side="right")

        # Parts table
        tbl_frame = tk.Frame(right, bg=t["BG"], highlightthickness=1,
                             highlightbackground=t["BORDER"])
        tbl_frame.pack(fill="both", expand=True)

        self.parts_tree = ttk.Treeview(
            tbl_frame, columns=self.PART_COLS, show="headings",
            style="PT.Treeview", selectmode="browse",
        )
        col_widths = [120, 160, 60, 180, 90, 110, 140]
        col_labels = ["Part Number", "Category", "Type", "Job", "Sub-Job",
                      "Catalog No", "Enclosure Size"]
        for col, label, w in zip(self.PART_COLS, col_labels, col_widths):
            self.parts_tree.heading(col, text=label)
            self.parts_tree.column(col, width=w, anchor="w",
                                   stretch=(col == "job"))
        py = ttk.Scrollbar(tbl_frame, orient="vertical", command=self.parts_tree.yview)
        self.parts_tree.configure(yscrollcommand=py.set)
        self.parts_tree.pack(side="left", fill="both", expand=True)
        py.pack(side="right", fill="y")

        # Archived rows get a warning background
        self.parts_tree.tag_configure("archived",
                                      background="#FFF7ED",
                                      foreground="#92400E")
        self.parts_tree.bind("<Double-1>", self._on_part_double_click)

        # Action row under table
        action = tk.Frame(right, bg=t["BG"])
        action.pack(fill="x", pady=(6, 0))
        tk.Button(
            action, text="Open in Explorer",
            bg=t["PANEL"], fg=t["TEXT"], font=t["BODY"],
            relief="flat", cursor="hand2",
            activebackground=t["BORDER"],
            command=self._open_selected,
        ).pack(side="right")
        tk.Label(action, text="Double-click a row or use the button to open the file location.",
                 bg=t["BG"], fg=t["SUBTLE"], font=t["SMALL"]).pack(side="left")

    # ── data ──
    def _on_search_changed(self):
        # Ignore placeholder text
        if self.job_search_var.get() == "Filter jobs…":
            return
        self._load_jobs()

    def _load_jobs(self):
        search = self.job_search_var.get().strip()
        if search == "Filter jobs…":
            search = ""
        try:
            all_jobs = self.db.get_jobs(search=search)
        except Exception:
            return

        self.job_tree.delete(*self.job_tree.get_children())
        self._job_index.clear()

        user_jobs: List[Tuple[str, dict]] = []
        for job in all_jobs:
            if job["part_count"] == 0:
                continue
            row = self.db.con.execute(
                "SELECT part_number FROM parts WHERE job_id=? AND category_code='003' LIMIT 1",
                (job["id"],)
            ).fetchone()
            sort_key = row["part_number"] if row else "003-99999"
            user_jobs.append((sort_key, dict(job)))

        user_jobs.sort(key=lambda x: x[0])

        for _, job in user_jobs:
            label = f"{job['sub_job']}   {job['job_name'] or ''}".strip()
            iid = self.job_tree.insert("", "end", text=label)
            self._job_index[iid] = job

        self.job_count_lbl.configure(text=f"{len(user_jobs)} job(s)")

        # If we had a job selected, try to re-select it
        if self._cur_job_id is not None:
            for iid, j in self._job_index.items():
                if j["id"] == self._cur_job_id:
                    self.job_tree.selection_set(iid)
                    self.job_tree.see(iid)
                    break

    def _on_job_select(self, _event):
        sel = self.job_tree.selection()
        if not sel:
            return
        job = self._job_index.get(sel[0])
        if not job:
            return
        self._cur_job_id = job["id"]
        self.parts_title_var.set(f"{job['sub_job']}  —  {job['job_name'] or ''}")

        bits = []
        if job.get("catalog_no"):     bits.append(f"Catalog: {job['catalog_no']}")
        if job.get("enclosure_size"): bits.append(f"Size: {job['enclosure_size']}")
        if job.get("is_archived"):    bits.append("⚠ Parts Archived")
        self.info_var.set("   ·   ".join(bits) if bits else "No PRF data found")
        self.info_lbl.configure(
            fg=self.theme["WARN"] if job.get("is_archived") else self.theme["SUBTLE"]
        )
        self._load_parts()

    def _load_parts(self):
        if self._cur_job_id is None:
            return
        # Category combo: "All Categories" → "", "100 – Subassembly" → "100"
        cat_text = self.cat_var.get()
        cat = "" if cat_text.startswith("All") else cat_text.split(" ")[0]
        ext_text = self.type_var.get()
        ext = "" if ext_text.startswith("All") else ext_text
        try:
            rows = self.db.get_parts(category=cat, job_id=self._cur_job_id, file_ext=ext)
        except Exception:
            return

        self.parts_tree.delete(*self.parts_tree.get_children())
        for p in rows:
            archived = any(part.lower() == "archive"
                           for part in Path(p["full_path"]).parts)
            job_str = f"{p['job_number']} {p['job_name'] or ''}".strip()
            values = (
                p["part_number"],
                f"{p['category_code']} – {p['category_name']}",
                p["file_ext"].upper(),
                job_str,
                p["sub_job"] or "—",
                p["catalog_no"] or "—",
                p["enclosure_size"] or "—",
            )
            tags = ("archived",) if archived else ()
            self.parts_tree.insert("", "end", values=values,
                                   tags=tags, iid=p["full_path"])
        self.part_count_lbl.configure(text=f"{len(rows)} part(s)")

    def _on_part_double_click(self, _event):
        self._open_selected()

    def _open_selected(self):
        sel = self.parts_tree.selection()
        if not sel:
            return
        open_path(sel[0])

    # ── public ──
    def refresh(self, user_prefix: Optional[str] = None):
        if user_prefix:
            self.user_prefix = user_prefix
        self._load_jobs()
        if self._cur_job_id is not None:
            self._load_parts()


# ═════════════════════════════════════════════════════════════════════
#  NEXT NUMBERS TAB
# ═════════════════════════════════════════════════════════════════════

class NextNumbersTab(tk.Frame):
    """Card grid showing latest/next part number per category, plus gap + dup panels."""

    def __init__(self, parent, db: Database, user_prefix: str,
                 cat_prefixes: Dict[str, str], theme: dict):
        super().__init__(parent, bg=theme["BG"])
        self.db = db
        self.user_prefix = user_prefix
        self.cat_prefixes = cat_prefixes or {}
        self.theme = theme

        self._cards: Dict[str, dict] = {}
        self._gap_rows: Dict[str, dict] = {}
        self._cached_gaps: Optional[Dict[str, List[str]]] = None
        self._reserved_numbers: Dict[str, set] = {}
        self._gap_worker: Optional[GapScanWorker] = None
        self._gap_queue: queue.Queue = queue.Queue()
        self._poll_job: Optional[str] = None
        self._gap_poll_job: Optional[str] = None

        self._build()
        self._schedule_poll()
        self._schedule_gap_poll()
        if self.user_prefix:
            self.after(300, self._start_gap_scan)

    # ── layout ──
    def _build(self):
        t = self.theme

        # Scrollable canvas so the whole content (cards + gap + dup) is reachable.
        wrapper = tk.Frame(self, bg=t["BG"])
        wrapper.pack(fill="both", expand=True)

        self._canvas = tk.Canvas(wrapper, bg=t["BG"], highlightthickness=0, bd=0)
        vsb = ttk.Scrollbar(wrapper, orient="vertical", command=self._canvas.yview)
        self._canvas.configure(yscrollcommand=vsb.set)
        self._canvas.pack(side="left", fill="both", expand=True)
        vsb.pack(side="right", fill="y")

        self._body = tk.Frame(self._canvas, bg=t["BG"])
        self._body_win = self._canvas.create_window((0, 0), window=self._body, anchor="nw")

        def _on_configure(_e):
            self._canvas.configure(scrollregion=self._canvas.bbox("all"))
        self._body.bind("<Configure>", _on_configure)

        def _on_canvas_configure(e):
            self._canvas.itemconfigure(self._body_win, width=e.width)
        self._canvas.bind("<Configure>", _on_canvas_configure)

        # Mouse wheel scrolling over the whole tab
        def _on_wheel(e):
            self._canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        self._canvas.bind_all("<MouseWheel>", _on_wheel, add="+")

        body = self._body

        # Header
        hdr = tk.Frame(body, bg=t["BG"])
        hdr.pack(fill="x", padx=16, pady=(14, 6))
        tk.Label(hdr, text="Next Part Numbers",
                 bg=t["BG"], fg=t["ACCENT"],
                 font=("Segoe UI", 12, "bold"), anchor="w").pack(side="left")
        tk.Label(hdr, text="Your next available number per category",
                 bg=t["BG"], fg=t["SUBTLE"], font=t["SMALL"]).pack(side="right")
        tk.Frame(body, bg=t["BORDER"], height=1).pack(fill="x", padx=16)

        # Card grid (3 columns)
        grid = tk.Frame(body, bg=t["BG"])
        grid.pack(fill="x", padx=16, pady=12)
        for col in range(3):
            grid.columnconfigure(col, weight=1, uniform="cards")

        categories = [(k, v) for k, v in CATEGORIES.items() if k != "003"]
        for idx, (code, name) in enumerate(categories):
            color = CAT_COLORS.get(code, t["ACCENT"])
            card = self._make_card(grid, code, name, color)
            card.grid(row=idx // 3, column=idx % 3,
                      padx=6, pady=6, sticky="nsew")

        # Gap Analysis panel
        self._gap_panel = self._build_gap_section(body)
        self._gap_panel.pack(fill="x", padx=16, pady=(0, 12))

        # Duplicate Parts panel
        self._dup_panel = self._build_duplicate_section(body)
        self._dup_panel.pack(fill="x", padx=16, pady=(0, 16))

    def _make_card(self, parent, code: str, name: str, color: str) -> tk.Frame:
        t = self.theme
        # Outer frame = colored left border via pack of a 4-px colored stripe.
        outer = tk.Frame(parent, bg=t["BORDER"], highlightthickness=1,
                         highlightbackground=t["BORDER"])
        stripe = tk.Frame(outer, bg=color, width=4)
        stripe.pack(side="left", fill="y")
        inner = tk.Frame(outer, bg=t["PANEL"])
        inner.pack(side="left", fill="both", expand=True)

        tk.Label(inner, text=name, bg=t["PANEL"], fg=color,
                 font=("Segoe UI", 11, "bold"), anchor="w").pack(
                 fill="x", padx=14, pady=(12, 0))
        tk.Label(inner, text=f"Category {code}", bg=t["PANEL"], fg=t["SUBTLE"],
                 font=t["SMALL"], anchor="w").pack(fill="x", padx=14)

        # Latest row
        latest_row = tk.Frame(inner, bg=t["PANEL"])
        latest_row.pack(fill="x", padx=14, pady=(10, 2))
        tk.Label(latest_row, text="Latest:", bg=t["PANEL"], fg=t["SUBTLE"],
                 font=t["SMALL"], width=7, anchor="w").pack(side="left")
        latest_val = tk.Label(latest_row, text="—", bg=t["PANEL"], fg=t["TEXT"],
                              font=("Segoe UI", 10, "bold"), anchor="w")
        latest_val.pack(side="left")

        # Next row
        next_row = tk.Frame(inner, bg=t["PANEL"])
        next_row.pack(fill="x", padx=14, pady=(2, 12))
        tk.Label(next_row, text="Next:", bg=t["PANEL"], fg=t["SUBTLE"],
                 font=t["SMALL"], width=7, anchor="w").pack(side="left")
        next_val = tk.Label(next_row, text="—", bg=t["PANEL"], fg=color,
                            font=("Segoe UI", 13, "bold"), anchor="w")
        next_val.pack(side="left", padx=(0, 6))

        gap_badge = tk.Label(next_row, text="GAP", bg="#F9E2AF", fg="#0A0A0A",
                             font=("Segoe UI", 8, "bold"),
                             padx=4, pady=0)
        # packed/unpacked dynamically

        copy_btn = tk.Button(
            next_row, text="Copy",
            bg=t["PANEL"], fg=color, font=t["SMALL"],
            activebackground=color, activeforeground="white",
            relief="solid", bd=1,
            highlightthickness=0, cursor="hand2",
            command=lambda: self._copy_next(code),
        )
        copy_btn.pack(side="right")

        self._cards[code] = {
            "code":       code,
            "color":      color,
            "latest_lbl": latest_val,
            "next_lbl":   next_val,
            "gap_badge":  gap_badge,
            "next_row":   next_row,
            "copy_btn":   copy_btn,
        }
        return outer

    def _build_gap_section(self, parent) -> tk.Frame:
        t = self.theme
        panel = tk.Frame(parent, bg=t["PANEL"], highlightthickness=1,
                         highlightbackground=t["BORDER"])

        hdr = tk.Frame(panel, bg=t["PANEL"])
        hdr.pack(fill="x", padx=14, pady=(12, 6))
        tk.Label(hdr, text="Gap Analysis", bg=t["PANEL"], fg=t["WARN"],
                 font=("Segoe UI", 11, "bold")).pack(side="left")
        self._gap_status_var = tk.StringVar(value="Pending — scan to check")
        tk.Label(hdr, textvariable=self._gap_status_var, bg=t["PANEL"],
                 fg=t["SUBTLE"], font=t["SMALL"]).pack(side="left", padx=(10, 0))

        self._gap_scan_btn = tk.Button(
            hdr, text="Scan Gaps",
            bg=t["WARN"], fg="white",
            activebackground="#B45309", activeforeground="white",
            font=("Segoe UI", 9, "bold"),
            relief="flat", cursor="hand2",
            command=self._start_gap_scan,
        )
        self._gap_scan_btn.pack(side="right")

        tk.Frame(panel, bg=t["BORDER"], height=1).pack(fill="x", padx=14)

        rows_frame = tk.Frame(panel, bg=t["PANEL"])
        rows_frame.pack(fill="x", padx=14, pady=(6, 12))

        for code, name in CATEGORIES.items():
            if code == "003":
                continue
            self._make_gap_row(rows_frame, code, name, CAT_COLORS.get(code, t["ACCENT"]))

        return panel

    def _make_gap_row(self, parent, code: str, name: str, color: str):
        t = self.theme
        container = tk.Frame(parent, bg=t["PANEL"])
        # container is NOT packed initially — revealed when gap scan completes

        toggle = tk.Button(
            container, text=f"▶  {name}  ({code})",
            bg=t["PANEL"], fg=color, font=t["SMALL"],
            relief="flat", cursor="hand2", anchor="w",
            activebackground=t["PANEL"],
        )
        toggle.pack(fill="x")

        detail_var = tk.StringVar(value="")
        detail = tk.Label(container, textvariable=detail_var, bg=t["PANEL"],
                          fg=t["SUBTLE"], font=("Consolas", 9),
                          wraplength=900, justify="left", anchor="w")
        # detail packed/unpacked by toggle

        def _on_toggle():
            row = self._gap_rows[code]
            row["open"] = not row["open"]
            if row["open"]:
                detail.pack(fill="x", padx=(24, 4), pady=(0, 6))
                txt = toggle.cget("text").replace("▶", "▼", 1)
                toggle.configure(text=txt)
            else:
                detail.pack_forget()
                txt = toggle.cget("text").replace("▼", "▶", 1)
                toggle.configure(text=txt)

        toggle.configure(command=_on_toggle)

        self._gap_rows[code] = {
            "container":  container,
            "toggle":     toggle,
            "detail":     detail,
            "detail_var": detail_var,
            "color":      color,
            "name":       name,
            "open":       False,
        }

    def _build_duplicate_section(self, parent) -> tk.Frame:
        t = self.theme
        panel = tk.Frame(parent, bg=t["PANEL"], highlightthickness=1,
                         highlightbackground=t["BORDER"])

        hdr = tk.Frame(panel, bg=t["PANEL"])
        hdr.pack(fill="x", padx=14, pady=(12, 6))
        tk.Label(hdr, text="Duplicate Parts", bg=t["PANEL"], fg=t["ERROR"],
                 font=("Segoe UI", 11, "bold")).pack(side="left")
        self._dup_status_var = tk.StringVar(value="Pending — click Scan to run")
        tk.Label(hdr, textvariable=self._dup_status_var, bg=t["PANEL"],
                 fg=t["SUBTLE"], font=t["SMALL"]).pack(side="left", padx=(10, 0))

        self._dup_scan_btn = tk.Button(
            hdr, text="Scan Duplicates",
            bg=t["ERROR"], fg="white",
            activebackground="#B91C1C", activeforeground="white",
            font=("Segoe UI", 9, "bold"),
            relief="flat", cursor="hand2",
            command=self._scan_duplicates_now,
        )
        self._dup_scan_btn.pack(side="right")

        tk.Frame(panel, bg=t["BORDER"], height=1).pack(fill="x", padx=14)

        body = tk.Frame(panel, bg=t["PANEL"])
        body.pack(fill="x", padx=14, pady=(6, 12))

        tree_frame = tk.Frame(body, bg=t["PANEL"], highlightthickness=1,
                              highlightbackground=t["BORDER"])
        tree_frame.pack(fill="both", expand=True)

        self._dup_tree = ttk.Treeview(
            tree_frame, columns=("job",), show="tree headings",
            style="PT.Treeview", height=8,
        )
        self._dup_tree.heading("#0", text="Part / Path")
        self._dup_tree.heading("job", text="Job")
        self._dup_tree.column("#0", width=520, anchor="w", stretch=True)
        self._dup_tree.column("job", width=180, anchor="w")
        self._dup_tree.tag_configure("group", foreground="#B91C1C")
        self._dup_tree.tag_configure("file", foreground=t["SUBTLE"])

        dy = ttk.Scrollbar(tree_frame, orient="vertical", command=self._dup_tree.yview)
        self._dup_tree.configure(yscrollcommand=dy.set)
        self._dup_tree.pack(side="left", fill="both", expand=True)
        dy.pack(side="right", fill="y")
        self._dup_tree.bind("<Double-1>", self._on_dup_double_click)

        self._dup_tree_frame = tree_frame

        self._dup_empty_var = tk.StringVar(value="")
        self._dup_empty_lbl = tk.Label(body, textvariable=self._dup_empty_var,
                                        bg=t["PANEL"], fg=t["SUBTLE"],
                                        font=t["SMALL"], anchor="w", pady=6)
        # Empty label packed/unpacked as needed. Tree hidden by default.
        self._dup_tree_frame.pack_forget()

        self._dup_body = body
        return panel

    # ── copy / refresh logic ──
    def _copy_next(self, code: str):
        card = self._cards.get(code)
        if not card:
            return
        text = card["next_lbl"].cget("text")
        if not text or text == "—":
            return

        # Click-time validation: ask Everything if this number is already on disk.
        # Catches stale-cache bugs where the gap list (or latest+1) collides with
        # files added since the last scan, by another engineer or in another session.
        conflict_path = is_part_number_taken(text)
        if conflict_path:
            messagebox.showwarning(
                "Number Already In Use",
                f"{text} is already on disk:\n\n{conflict_path}\n\n"
                "The suggestion was stale. Refreshing the gap cache now — "
                "click Copy again for a fresh number.",
                parent=self,
            )
            self._reserved_numbers.setdefault(code, set()).add(text)
            if self._cached_gaps and code in self._cached_gaps:
                try:
                    self._cached_gaps[code].remove(text)
                except ValueError:
                    pass
                if not self._cached_gaps[code]:
                    del self._cached_gaps[code]
            self.refresh()
            self._start_gap_scan()
            return

        num_only = text.split("-", 1)[1] if "-" in text else text
        try:
            self.clipboard_clear()
            self.clipboard_append(num_only)
        except tk.TclError:
            pass
        if self._cached_gaps and code in self._cached_gaps:
            try:
                self._cached_gaps[code].remove(text)
            except ValueError:
                pass
            if not self._cached_gaps[code]:
                del self._cached_gaps[code]
        self._reserved_numbers.setdefault(code, set()).add(text)
        self.refresh()

    def clear_reservations(self):
        self._reserved_numbers.clear()

    def _safe_refresh(self):
        try:
            self.refresh()
        except Exception:
            pass

    def _schedule_poll(self):
        self._poll_job = self.after(DB_POLL_MS, self._tick_poll)

    def _tick_poll(self):
        self._safe_refresh()
        self._schedule_poll()

    def _schedule_gap_poll(self):
        """Drain GapScanWorker results from queue on the main thread."""
        try:
            while True:
                tag, payload = self._gap_queue.get_nowait()
                if tag == "gaps:done":
                    self._on_gap_scan_done(payload)
        except queue.Empty:
            pass
        self._gap_poll_job = self.after(150, self._schedule_gap_poll)

    def _start_gap_scan(self):
        if self._gap_worker and self._gap_worker.is_alive():
            return
        if not self.user_prefix:
            return
        self._gap_status_var.set("Scanning Everything…")
        self._gap_scan_btn.configure(state="disabled")
        self._gap_worker = GapScanWorker(self.user_prefix, self.cat_prefixes,
                                          self._gap_queue)
        self._gap_worker.start()

    def _on_gap_scan_done(self, result: Dict[str, List[str]]):
        self._cached_gaps = result
        self._gap_worker = None
        try:
            self._gap_scan_btn.configure(state="normal")
        except tk.TclError:
            return
        self._reserved_numbers.clear()
        self.refresh()

    def _scan_duplicates_now(self):
        if not self.user_prefix:
            self._dup_status_var.set("No user configured")
            return
        self._dup_scan_btn.configure(state="disabled")
        self._dup_status_var.set("Scanning…")
        try:
            groups = self.db.get_duplicate_parts(self.user_prefix)
        except Exception as e:
            self._dup_status_var.set(f"Error: {e}")
            self._dup_scan_btn.configure(state="normal")
            return

        for iid in self._dup_tree.get_children():
            self._dup_tree.delete(iid)

        if not groups:
            self._dup_tree_frame.pack_forget()
            self._dup_empty_var.set("No duplicate parts found. Your part numbers are all unique.")
            self._dup_empty_lbl.pack(fill="x")
            self._dup_status_var.set("0 duplicates")
        else:
            self._dup_empty_lbl.pack_forget()
            self._dup_tree_frame.pack(fill="both", expand=True)
            total_files = 0
            for g in groups:
                pn   = g["part_number"]
                ext  = g["file_ext"]
                kids = g["entries"]
                total_files += len(kids)
                top_iid = self._dup_tree.insert(
                    "", "end",
                    text=f"{pn}.{ext.lstrip('.').lower()}  ({len(kids)} copies)",
                    values=("",), tags=("group",), open=True,
                )
                for entry in kids:
                    job_label = ""
                    if entry.get("job_number"):
                        job_label = entry["job_number"]
                        if entry.get("sub_job") and entry["sub_job"] != entry["job_number"]:
                            job_label += f" / {entry['sub_job']}"
                    self._dup_tree.insert(
                        top_iid, "end",
                        text=entry["full_path"],
                        values=(job_label,), tags=("file",),
                        iid=f"dup|{top_iid}|{entry['full_path']}",
                    )
            n = len(groups)
            self._dup_status_var.set(
                f"{n} duplicated part number{'s' if n != 1 else ''} ({total_files} files total)"
            )
        self._dup_scan_btn.configure(state="normal")

    def _on_dup_double_click(self, _event):
        sel = self._dup_tree.selection()
        if not sel:
            return
        iid = sel[0]
        if iid.startswith("dup|"):
            # iid format: "dup|<group>|<full_path>"
            _, _, path = iid.split("|", 2)
            open_path(path)

    # ── refresh (cards + gap panel) ──
    def refresh(self, user_prefix: Optional[str] = None,
                cat_prefixes: Optional[Dict[str, str]] = None):
        if user_prefix:
            self.user_prefix = user_prefix
        if cat_prefixes is not None:
            self.cat_prefixes = cat_prefixes

        rows = self.db.latest_by_category(self.user_prefix, self.cat_prefixes)
        existing = {r["category_code"]: r for r in rows}
        gaps = self._cached_gaps if self._cached_gaps is not None else {}

        for code, card in self._cards.items():
            row      = existing.get(code)
            latest   = row["latest_part"] if row else ""
            eff_pfx  = self.cat_prefixes.get(code, self.user_prefix)
            gap_list = gaps.get(code, [])
            reserved = self._reserved_numbers.get(code, set())

            available_gaps = [g for g in gap_list if g not in reserved]
            if available_gaps:
                nxt = available_gaps[0]
                is_gap = True
            else:
                nxt = _next_part(eff_pfx or "", code, latest)
                safety = 0
                while nxt in reserved and safety < 10_000:
                    five = nxt.split("-")[1]
                    nxt  = f"{code}-{str(int(five) + 1).zfill(5)}"
                    safety += 1
                is_gap = False

            card["latest_lbl"].configure(text=latest if latest else "None yet")
            card["next_lbl"].configure(text=nxt)

            badge = card["gap_badge"]
            if is_gap:
                if not badge.winfo_ismapped():
                    badge.pack(side="left", padx=(0, 6),
                               before=card["copy_btn"])
            else:
                if badge.winfo_ismapped():
                    badge.pack_forget()

        # Gap panel rendering
        scanned = self._cached_gaps is not None
        cats_with_gaps = 0
        for code, row_info in self._gap_rows.items():
            container = row_info["container"]
            toggle    = row_info["toggle"]
            detail    = row_info["detail"]
            name      = row_info["name"]
            color     = row_info["color"]
            gap_list  = gaps.get(code, [])

            if not scanned:
                container.pack_forget()
                continue

            if not container.winfo_ismapped():
                container.pack(fill="x")
            arrow = "▼" if row_info["open"] else "▶"
            if gap_list:
                cats_with_gaps += 1
                n = len(gap_list)
                label = "gap" if n == 1 else "gaps"
                toggle.configure(
                    text=f"{arrow}  {name}  ({code})   —   {n} {label} missing",
                    fg=color,
                )
                row_info["detail_var"].set("  ".join(gap_list))
            else:
                toggle.configure(
                    text=f"   ✓  {name}  ({code})   —   no gaps",
                    fg=self.theme["SUBTLE"],
                )
                row_info["detail_var"].set("")
                if row_info["open"]:
                    # auto-collapse empties so the panel stays tidy
                    detail.pack_forget()
                    row_info["open"] = False

        if scanned:
            if cats_with_gaps:
                word = "category" if cats_with_gaps == 1 else "categories"
                self._gap_status_var.set(f"{cats_with_gaps} {word} with gaps")
            else:
                self._gap_status_var.set("All clear — no gaps found")
        else:
            self._gap_status_var.set("Pending — scan to check")

    # ── lifecycle ──
    def destroy(self):
        for jobref in (self._poll_job, self._gap_poll_job):
            if jobref is not None:
                try:
                    self.after_cancel(jobref)
                except Exception:
                    pass
        super().destroy()


# ═════════════════════════════════════════════════════════════════════
#  PARTS TRACKER PANEL (top-level wrapper)
# ═════════════════════════════════════════════════════════════════════

class PartsTrackerPanel(tk.Frame):
    """
    Top-level embedded panel with:
      • Header bar: user label + Change User / Clear All / Refresh buttons
      • ttk.Notebook: My Parts + Next Numbers tabs
    """

    SUBTITLE = ("Tracks SolidWorks parts across jobs via Everything + PRF data. "
                "Use My Parts to browse by job, and Next Numbers to grab the "
                "next available part ID per category.")

    def __init__(self, parent, theme: Optional[dict] = None):
        # Theme dict expected keys: BG, PANEL, BORDER, ACCENT, ACCENT_H,
        # TEXT, SUBTLE, SUCCESS, ERROR, WARN, BODY, SMALL
        self.theme = theme or DEFAULT_THEME
        super().__init__(parent, bg=self.theme["BG"])

        self.db = Database()
        self.user_name = os.environ.get("USERNAME", "User")
        self.user_prefix = self.db.get("user_prefix", "") or ""
        self.cat_prefixes = self.db.get_cat_prefixes() or {}

        self._scan_queue: queue.Queue = queue.Queue()
        self._scan_worker: Optional[ScanWorker] = None
        self._scan_dlg: Optional[ScanDialog] = None
        self._scan_silent = False
        self._scan_poll_job: Optional[str] = None
        self._rescan_job: Optional[str] = None

        self._tab_my: Optional[MyPartsTab] = None
        self._tab_next: Optional[NextNumbersTab] = None

        self._configure_ttk_styles()
        self._build()
        self._schedule_scan_poll()

        # First-run setup vs kick-off scan
        if not self.user_prefix:
            self.after(120, self._first_run_setup)
        else:
            self.after(250, self._start_scan)
            self._schedule_auto_rescan()

    # ── ttk styling shared by the whole panel ──
    def _configure_ttk_styles(self):
        t = self.theme
        s = ttk.Style(self)
        try:
            base = s.theme_use()
            _ = base
        except Exception:
            pass

        s.configure("PT.TNotebook", background=t["BG"], borderwidth=0)
        s.configure("PT.TNotebook.Tab",
                    background=t["BG"], foreground=t["SUBTLE"],
                    padding=[14, 6], font=t["BODY"])
        s.map("PT.TNotebook.Tab",
              background=[("selected", t["PANEL"])],
              foreground=[("selected", t["ACCENT"])])

        s.configure("PT.Treeview",
                    background=t["PANEL"],
                    foreground=t["TEXT"],
                    fieldbackground=t["PANEL"],
                    rowheight=24, font=t["BODY"], borderwidth=0)
        s.configure("PT.Treeview.Heading",
                    background=t["BG"], foreground=t["TEXT"],
                    font=("Segoe UI", 9, "bold"), relief="flat")
        s.map("PT.Treeview",
              background=[("selected", t["ACCENT"])],
              foreground=[("selected", "white")])

        s.configure("PT.TCombobox", font=t["BODY"])
        s.configure("PT.TPanedwindow", background=t["BG"])
        s.configure("PT.Horizontal.TProgressbar",
                    troughcolor=t["BORDER"], background=t["ACCENT"],
                    thickness=8)

    # ── UI ──
    def _build(self):
        t = self.theme

        # Header bar — user label + actions
        bar = tk.Frame(self, bg=t["BG"])
        bar.pack(fill="x", pady=(0, 10))

        self.user_var = tk.StringVar(value=self._user_label_text())
        tk.Label(bar, textvariable=self.user_var, bg=t["BG"], fg=t["SUBTLE"],
                 font=t["BODY"]).pack(side="left")

        btn_row = tk.Frame(bar, bg=t["BG"])
        btn_row.pack(side="right")

        def _btn(text, cmd, primary=False, danger=False):
            if primary:
                bg, fg, hover = t["ACCENT"], "white", t["ACCENT_H"]
            elif danger:
                bg, fg, hover = t["ERROR"], "white", "#B91C1C"
            else:
                bg, fg, hover = t["PANEL"], t["TEXT"], t["BORDER"]
            b = tk.Button(btn_row, text=text, bg=bg, fg=fg,
                          activebackground=hover, activeforeground=fg if primary or danger else t["TEXT"],
                          font=t["BODY"], relief="flat", cursor="hand2",
                          padx=12, pady=4, command=cmd)
            b.pack(side="left", padx=(6, 0))
            return b

        _btn("Change User", self._change_user)
        _btn("Clear All",   self._clear_all, danger=True)
        self.refresh_btn = _btn("Refresh", self._start_scan, primary=True)

        # Status line (shown briefly after scans)
        self.status_var = tk.StringVar(value="")
        tk.Label(self, textvariable=self.status_var, bg=t["BG"],
                 fg=t["SUCCESS"], font=t["SMALL"], anchor="w").pack(
                 fill="x", pady=(0, 4))

        # Notebook
        self.nb = ttk.Notebook(self, style="PT.TNotebook")
        self.nb.pack(fill="both", expand=True)

        self._render_tabs()

    def _render_tabs(self):
        t = self.theme
        # Remove existing tabs (when user changes)
        for tab_id in list(self.nb.tabs()):
            self.nb.forget(tab_id)
        self._tab_my = None
        self._tab_next = None

        if not self.user_prefix:
            placeholder = tk.Frame(self.nb, bg=t["BG"])
            tk.Label(placeholder,
                     text="Click 'Change User' to configure your SolidWorks user ID.",
                     bg=t["BG"], fg=t["SUBTLE"], font=t["BODY"]).pack(pady=60)
            self.nb.add(placeholder, text="  Setup  ")
            return

        self._tab_my = MyPartsTab(self.nb, self.db, self.user_prefix, t)
        self._tab_next = NextNumbersTab(self.nb, self.db, self.user_prefix,
                                         self.cat_prefixes, t)
        self.nb.add(self._tab_my, text=f"  My Parts ({self.user_prefix})  ")
        self.nb.add(self._tab_next, text="  Next Numbers  ")

    def _user_label_text(self) -> str:
        if not self.user_prefix:
            return "No user configured — click 'Change User' to set one up."
        return f"{self.user_name}  ·  prefix  {self.user_prefix}"

    # ── setup / change user ──
    def _first_run_setup(self):
        dlg = SetupDialog(self, self.theme,
                          prefill_id=self.user_prefix,
                          prefill_cats=self.cat_prefixes)
        self.wait_window(dlg)
        if not dlg.result:
            return
        v = dlg.result
        self.user_prefix = v["user_prefix"]
        self.cat_prefixes = v.get("cat_prefixes", {}) or {}
        self.db.put("user_id", v["user_id"])
        self.db.put("user_prefix", self.user_prefix)
        self.db.set_cat_prefixes(self.cat_prefixes)
        self.user_var.set(self._user_label_text())
        self._render_tabs()
        self.after(200, self._start_scan)
        self._schedule_auto_rescan()

    def _change_user(self):
        old_prefix = self.user_prefix
        old_cats = dict(self.cat_prefixes)
        dlg = SetupDialog(self, self.theme,
                          prefill_id=self.user_prefix,
                          prefill_cats=self.cat_prefixes)
        self.wait_window(dlg)
        if not dlg.result:
            return
        v = dlg.result
        new_prefix = v["user_prefix"]
        new_cats = v.get("cat_prefixes", {}) or {}
        prefix_changed = (new_prefix != old_prefix) or (new_cats != old_cats)

        self.user_prefix = new_prefix
        self.cat_prefixes = new_cats
        self.db.put("user_id", v["user_id"])
        self.db.put("user_prefix", self.user_prefix)
        self.db.set_cat_prefixes(self.cat_prefixes)
        self.user_var.set(self._user_label_text())

        if prefix_changed:
            self.db.clear_all()

        self._render_tabs()
        self._start_scan()

    # ── scanning ──
    def _start_scan(self, silent: bool = False):
        if not self.user_prefix:
            return
        if self._scan_worker and self._scan_worker.is_alive():
            return
        self._scan_silent = silent
        if not silent:
            self._scan_dlg = ScanDialog(self, self.theme, on_cancel=self._cancel_scan)
            try:
                self.refresh_btn.configure(state="disabled")
            except tk.TclError:
                pass
        self._scan_worker = ScanWorker(self.db, self.user_prefix, self._scan_queue)
        self._scan_worker.start()

    def _cancel_scan(self):
        if self._scan_worker and self._scan_worker.is_alive():
            self._scan_worker.cancel()

    def _schedule_scan_poll(self):
        try:
            while True:
                tag, payload = self._scan_queue.get_nowait()
                if tag == "scan:progress":
                    pct, msg = payload
                    if self._scan_dlg and not self._scan_silent:
                        self._scan_dlg.update_progress(pct, msg)
                elif tag == "scan:folders":
                    pass  # watchdog not implemented; auto-rescan covers it
                elif tag == "scan:done":
                    new_parts, new_jobs = payload
                    self._finish_scan(success=True, new_parts=new_parts, new_jobs=new_jobs)
                elif tag == "scan:err":
                    self._finish_scan(success=False, err=payload)
                elif tag == "scan:cancel":
                    self._finish_scan(success=False, cancelled=True)
        except queue.Empty:
            pass
        self._scan_poll_job = self.after(100, self._schedule_scan_poll)

    def _finish_scan(self, success: bool,
                     new_parts: int = 0, new_jobs: int = 0,
                     err: str = "", cancelled: bool = False):
        if self._scan_dlg:
            try:
                self._scan_dlg.destroy()
            except tk.TclError:
                pass
            self._scan_dlg = None
        try:
            self.refresh_btn.configure(state="normal")
        except tk.TclError:
            return
        self._scan_worker = None

        if success:
            if not self._scan_silent:
                self.status_var.set(
                    f"Scan complete — {new_parts:,} part(s) processed, {new_jobs} new job(s)"
                )
                self.after(8000, lambda: self.status_var.set(""))
            self._reload_tabs()
        elif cancelled:
            if not self._scan_silent:
                self.status_var.set("Scan cancelled.")
                self.after(4000, lambda: self.status_var.set(""))
            # Still reload tabs — DB has partial data from the cancelled run
            self._reload_tabs()
        else:
            if not self._scan_silent:
                messagebox.showerror("Scan Error", err or "Unknown error", parent=self)

        self._scan_silent = False

    def _reload_tabs(self):
        if self._tab_my is not None:
            try:
                self._tab_my.refresh(self.user_prefix)
            except tk.TclError:
                pass
        if self._tab_next is not None:
            try:
                self._tab_next.clear_reservations()
                self._tab_next.refresh(self.user_prefix, self.cat_prefixes)
            except tk.TclError:
                pass
        # Keep tab title in sync with user prefix
        try:
            tabs = self.nb.tabs()
            if tabs:
                self.nb.tab(tabs[0], text=f"  My Parts ({self.user_prefix})  ")
        except tk.TclError:
            pass

    # ── auto-rescan (SMB watcher fallback) ──
    def _schedule_auto_rescan(self):
        if self._rescan_job is not None:
            try:
                self.after_cancel(self._rescan_job)
            except Exception:
                pass
        self._rescan_job = self.after(
            AUTO_RESCAN_MINUTES * 60 * 1000,
            self._tick_auto_rescan,
        )

    def _tick_auto_rescan(self):
        if self.user_prefix and not (self._scan_worker and self._scan_worker.is_alive()):
            self._start_scan(silent=True)
        self._schedule_auto_rescan()

    # ── clear all ──
    def _clear_all(self):
        ans = messagebox.askyesno(
            "Clear all scanned data?",
            "This wipes every scanned job and part from the local database.\n\n"
            "Your user ID stays; a fresh scan starts right after. Continue?",
            parent=self,
        )
        if not ans:
            return
        self.db.clear_all()
        self._reload_tabs()
        self._start_scan()

    # ── lifecycle ──
    def destroy(self):
        for jobref in (self._scan_poll_job, self._rescan_job):
            if jobref is not None:
                try:
                    self.after_cancel(jobref)
                except Exception:
                    pass
        try:
            self.db.close()
        except Exception:
            pass
        super().destroy()


# ═════════════════════════════════════════════════════════════════════
#  DEFAULT THEME (fallback if host doesn't pass one)
# ═════════════════════════════════════════════════════════════════════

DEFAULT_THEME: Dict[str, object] = {
    "BG":       "#F6F8FB",
    "PANEL":    "#FFFFFF",
    "BORDER":   "#D8DEE9",
    "ACCENT":   "#1F6FEB",
    "ACCENT_H": "#1558D6",
    "TEXT":     "#1F2937",
    "SUBTLE":   "#6B7280",
    "SUCCESS":  "#16A34A",
    "ERROR":    "#DC2626",
    "WARN":     "#D97706",
    "BODY":     ("Segoe UI", 10),
    "SMALL":    ("Segoe UI", 9),
}


# ═════════════════════════════════════════════════════════════════════
#  STANDALONE / SELF-TEST
# ═════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Parts Tracker (standalone)")
    root.geometry("1200x800")
    root.configure(bg=DEFAULT_THEME["BG"])
    panel = PartsTrackerPanel(root)
    panel.pack(fill="both", expand=True, padx=12, pady=12)
    root.mainloop()
